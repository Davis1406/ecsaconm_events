import io
import os
import re
import uuid
from typing import Annotated
from fastapi import APIRouter, BackgroundTasks, Depends, Form, HTTPException, status, Query, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.database import get_db
from dependencies.auth_dependency import Auth, get_current_user
from models.models import Abstract, AbstractAuthor, AbstractReviewer, AbstractReview, User, UserRole, Role, Registration, Event, AbstractStatus
from datetime import datetime, timezone
from sqlalchemy import text as _sql_text
from schemas.events_space import (
    AbstractSubmitSchema, AbstractUpdateSchema,
    AssignReviewerSchema, AbstractReviewSchema,
    CreateReviewerSchema,
)

PRESENTATION_UPLOAD_DIR = "uploads/presentations"
os.makedirs(PRESENTATION_UPLOAD_DIR, exist_ok=True)
ALLOWED_PRESENTATION_EXTS = {".ppt", ".pptx", ".pdf", ".zip", ".mp4"}
MAX_PRESENTATION_MB = 100

router = APIRouter()

user_dependency = Annotated[dict, Depends(get_current_user)]


def get_auth_dep(db: Session = Depends(get_db)) -> Auth:
    return Auth(db)


def _serialize_abstract(a: Abstract):
    return {
        "id": a.id,
        "event_id": a.event_id,
        "event": a.event.event if a.event else None,
        "title": a.title,
        "abstract_text": a.abstract_text,
        "keywords": a.keywords,
        "track": a.track,
        "presentation_type": a.presentation_type.value if a.presentation_type else None,
        "status": a.status.value if a.status else None,
        "word_count": a.word_count,
        "submitted_by": a.submitted_by,
        "submitter_name": f"{a.submitter.firstname} {a.submitter.lastname}" if a.submitter else None,
        "submitter_email": a.submitter.email if a.submitter else None,
        "created_at": a.created_at,
        "updated_at": a.updated_at,
        "authors": [
            {
                "id": au.id,
                "firstname": au.firstname,
                "lastname": au.lastname,
                "email": au.email,
                "affiliation": au.affiliation,
                "country": au.country,
                "is_presenting": au.is_presenting,
                "author_order": au.author_order,
            }
            for au in a.authors
        ],
        "reviewer_assignments": [
            {
                "id": ra.id,
                "reviewer_id": ra.reviewer_id,
                "reviewer_name": f"{ra.reviewer.firstname} {ra.reviewer.lastname}" if ra.reviewer else None,
                "reviewer_email": ra.reviewer.email if ra.reviewer else None,
                "assigned_at": ra.assigned_at,
                "completed": ra.completed,
                "review": {
                    "id": ra.review.id,
                    "relevance_score": ra.review.relevance_score,
                    "methodology_score": ra.review.methodology_score,
                    "originality_score": ra.review.originality_score,
                    "overall_score": ra.review.overall_score,
                    "recommendation": ra.review.recommendation.value,
                    "comments": ra.review.comments,
                    "submitted_at": ra.review.submitted_at,
                } if ra.review else None,
            }
            for ra in a.reviewer_assignments
        ],
    }


@router.post("/", status_code=status.HTTP_201_CREATED)
def submit_abstract(
    schema: AbstractSubmitSchema,
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    word_count = len(schema.abstract_text.split())
    abstract = Abstract(
        event_id=schema.event_id,
        submitted_by=current_user["user_id"],
        title=schema.title,
        abstract_text=schema.abstract_text,
        keywords=schema.keywords,
        track=schema.track,
        presentation_type=schema.presentation_type,
        word_count=word_count,
    )
    db.add(abstract)
    db.flush()

    for i, author in enumerate(schema.authors):
        db.add(AbstractAuthor(
            abstract_id=abstract.id,
            firstname=author.firstname,
            lastname=author.lastname,
            email=author.email,
            affiliation=author.affiliation,
            country=author.country,
            is_presenting=author.is_presenting,
            author_order=i,
        ))

    db.commit()
    db.refresh(abstract)
    # reload with relationships
    abstract = db.query(Abstract).options(
        joinedload(Abstract.authors),
        joinedload(Abstract.submitter),
        joinedload(Abstract.event),
        joinedload(Abstract.reviewer_assignments),
    ).filter(Abstract.id == abstract.id).first()
    return _serialize_abstract(abstract)


@router.get("/stats")
def abstract_stats(
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
    event_id: int = Query(None),
):
    """Summary counts for the Accepted Abstracts dashboard."""
    from sqlalchemy import func
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])

    base = db.query(Abstract).filter(
        Abstract.deleted_at == None,
        Abstract.status == "accepted",
    )
    if event_id:
        base = base.filter(Abstract.event_id == event_id)

    total     = base.count()
    oral      = base.filter(Abstract.presentation_type == "oral").count()
    poster    = base.filter(Abstract.presentation_type == "poster").count()

    # Unique presenters (by distinct email on is_presenting authors)
    presenter_q = (
        db.query(AbstractAuthor.email)
        .join(Abstract, AbstractAuthor.abstract_id == Abstract.id)
        .filter(
            Abstract.deleted_at == None,
            Abstract.status == "accepted",
            AbstractAuthor.is_presenting == True,
            AbstractAuthor.email != None,
            AbstractAuthor.email != "",
        )
    )
    if event_id:
        presenter_q = presenter_q.filter(Abstract.event_id == event_id)
    unique_presenter_emails = {r.email.strip().lower() for r in presenter_q.all()}
    unique_presenters = len(unique_presenter_emails)

    # Presenters with 2+ abstracts (by email)
    from sqlalchemy import func as sqlfunc
    multi_q = (
        db.query(AbstractAuthor.email, sqlfunc.count(AbstractAuthor.abstract_id).label("cnt"))
        .join(Abstract, AbstractAuthor.abstract_id == Abstract.id)
        .filter(
            Abstract.deleted_at == None,
            Abstract.status == "accepted",
            AbstractAuthor.is_presenting == True,
            AbstractAuthor.email != None,
            AbstractAuthor.email != "",
        )
        .group_by(AbstractAuthor.email)
        .having(sqlfunc.count(AbstractAuthor.abstract_id) > 1)
    )
    if event_id:
        multi_q = multi_q.filter(Abstract.event_id == event_id)
    multi_presenters = multi_q.count()

    # Registered / paid counts (by presenter email)
    all_pres_q = (
        db.query(AbstractAuthor.email)
        .join(Abstract, AbstractAuthor.abstract_id == Abstract.id)
        .filter(
            Abstract.deleted_at == None,
            Abstract.status == "accepted",
            AbstractAuthor.is_presenting == True,
            AbstractAuthor.email != None,
            AbstractAuthor.email != "",
        )
    )
    if event_id:
        all_pres_q = all_pres_q.filter(Abstract.event_id == event_id)

    registered_count = 0
    paid_count = 0
    seen_reg: set = set()
    for row in all_pres_q.all():
        email = row.email.strip().lower()
        if email in seen_reg:
            continue
        seen_reg.add(email)
        user = db.query(User).filter(User.email == email).first()
        if user:
            reg = db.query(Registration).filter(Registration.user_id == user.id).first()
            if reg:
                registered_count += 1
                if reg.paid:
                    paid_count += 1

    return {
        "total": total,
        "oral": oral,
        "poster": poster,
        "unique_presenters": unique_presenters,
        "multi_presenters": multi_presenters,
        "registered": registered_count,
        "paid": paid_count,
        "not_registered": unique_presenters - registered_count,
    }


@router.get("/")
def list_abstracts(
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
    event_id: int = None,
    skip: int = 0,
    limit: int = 25,
    search: str = Query(None),
    status: str = Query(None),
    presentation_type: str = Query(None),
    presenter_email: str = Query(None),
    sort_by: str = Query("created_at"),   # title | type | created_at
    sort_dir: str = Query("desc"),         # asc | desc
    presenter_registered: str = Query(None),  # "yes" | "no"
    presenter_paid: str = Query(None),        # "yes" | "no"
):
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])
    from sqlalchemy import or_
    q = db.query(Abstract).filter(Abstract.deleted_at == None)
    if event_id:
        q = q.filter(Abstract.event_id == event_id)
    if status:
        q = q.filter(Abstract.status == status)
    if presentation_type:
        q = q.filter(Abstract.presentation_type == presentation_type)
    if search:
        q = q.filter(or_(
            Abstract.title.ilike(f"%{search}%"),
            Abstract.keywords.ilike(f"%{search}%"),
            Abstract.track.ilike(f"%{search}%"),
        ))
    # Filter to abstracts where the presenter has 2+ accepted abstracts
    if presenter_email == "multi":
        from sqlalchemy import func as sqlfunc
        multi_emails = (
            db.query(AbstractAuthor.email)
            .join(Abstract, AbstractAuthor.abstract_id == Abstract.id)
            .filter(
                Abstract.deleted_at == None,
                Abstract.status == "accepted",
                AbstractAuthor.is_presenting == True,
                AbstractAuthor.email != None,
                AbstractAuthor.email != "",
            )
            .group_by(AbstractAuthor.email)
            .having(sqlfunc.count(AbstractAuthor.abstract_id) > 1)
            .subquery()
        )
        q = q.join(AbstractAuthor, AbstractAuthor.abstract_id == Abstract.id).filter(
            AbstractAuthor.is_presenting == True,
            AbstractAuthor.email.in_(db.query(multi_emails.c.email)),
        )
    # Filter by presenter registration / payment status
    if presenter_registered in ("yes", "no") or presenter_paid in ("yes", "no"):
        # Gather all presenting author emails for accepted abstracts in scope
        pres_q = (
            db.query(AbstractAuthor.email, AbstractAuthor.abstract_id)
            .join(Abstract, AbstractAuthor.abstract_id == Abstract.id)
            .filter(
                Abstract.deleted_at == None,
                Abstract.status == "accepted",
                AbstractAuthor.is_presenting == True,
                AbstractAuthor.email != None,
                AbstractAuthor.email != "",
            )
        )
        if event_id:
            pres_q = pres_q.filter(Abstract.event_id == event_id)

        # Build email → abstract_id map and evaluate registration per email
        matched_abstract_ids = set()
        seen_emails: dict = {}
        for row in pres_q.all():
            email = row.email.strip().lower()
            if email not in seen_emails:
                user = db.query(User).filter(User.email == email).first()
                has_registered = False
                has_paid = False
                if user:
                    reg = db.query(Registration).filter(
                        Registration.user_id == user.id,
                    ).first()
                    if reg:
                        has_registered = True
                        has_paid = bool(reg.paid)
                seen_emails[email] = {"has_registered": has_registered, "has_paid": has_paid}
            st = seen_emails[email]
            reg_ok = (
                presenter_registered is None or
                (presenter_registered == "yes" and st["has_registered"]) or
                (presenter_registered == "no" and not st["has_registered"])
            )
            paid_ok = (
                presenter_paid is None or
                (presenter_paid == "yes" and st["has_paid"]) or
                (presenter_paid == "no" and not st["has_paid"])
            )
            if reg_ok and paid_ok:
                matched_abstract_ids.add(row.abstract_id)

        q = q.filter(Abstract.id.in_(matched_abstract_ids))

    total = q.count()
    # Sort
    _sort_map = {
        "title":        Abstract.title,
        "type":         Abstract.presentation_type,
        "created_at":   Abstract.created_at,
    }
    sort_col = _sort_map.get(sort_by, Abstract.created_at)
    order_expr = sort_col.asc() if sort_dir == "asc" else sort_col.desc()
    abstracts = q.options(
        joinedload(Abstract.authors),
        joinedload(Abstract.submitter),
        joinedload(Abstract.event),
    ).order_by(order_expr).offset(skip).limit(limit).all()
    return {"data": [_serialize_abstract(a) for a in abstracts], "total": total}


@router.get("/export")
def export_abstracts(
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
    event_id: int = None,
    status_filter: str = Query(None, alias="status"),
    search: str = Query(None),
):
    auth_dependency.secure_access("EXPORT_ABSTRACTS", current_user["user_id"])
    from sqlalchemy import or_
    q = db.query(Abstract).options(
        joinedload(Abstract.authors),
        joinedload(Abstract.submitter),
        joinedload(Abstract.event),
        joinedload(Abstract.reviewer_assignments).joinedload(AbstractReviewer.reviewer),
        joinedload(Abstract.reviewer_assignments).joinedload(AbstractReviewer.review),
    ).filter(Abstract.deleted_at == None)
    if event_id:
        q = q.filter(Abstract.event_id == event_id)
    if status_filter:
        q = q.filter(Abstract.status == status_filter)
    if search:
        term = f"%{search}%"
        q = q.filter(or_(
            Abstract.title.ilike(term),
            Abstract.keywords.ilike(term),
        ))
    abstracts = q.order_by(Abstract.created_at.desc()).all()

    wb = Workbook()

    # ── Sheet 1: Abstracts ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Abstracts"

    header_fill = PatternFill("solid", start_color="0095B6")
    alt_fill    = PatternFill("solid", start_color="E8F4F8")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    body_font   = Font(name="Arial", size=10)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    headers = [
        "#", "Title", "Event", "Track", "Presentation Type", "Status",
        "Word Count", "Keywords", "Submitter Name", "Submitter Email",
        "First Author", "First Author Email", "First Author Affiliation",
        "First Author Country", "All Authors", "Reviewers Assigned",
        "Reviews Completed", "Avg Relevance", "Avg Methodology",
        "Avg Originality", "Avg Overall", "Date Submitted",
    ]

    ws.row_dimensions[1].height = 22
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    ws.freeze_panes = "A2"

    for ri, a in enumerate(abstracts, 2):
        ws.row_dimensions[ri].height = 16
        use_fill = alt_fill if ri % 2 == 0 else PatternFill("solid", start_color="FFFFFF")

        authors = sorted(a.authors, key=lambda x: x.author_order)
        first   = authors[0] if authors else None
        all_authors = "; ".join(
            f"{au.firstname} {au.lastname}" + (f" ({au.affiliation})" if au.affiliation else "")
            for au in authors
        )

        assignments = a.reviewer_assignments
        reviews = [r.review for r in assignments if r.review]
        avg = lambda field: round(sum(getattr(r, field) for r in reviews) / len(reviews), 1) if reviews else ""

        row = [
            a.id,
            a.title,
            a.event.event if a.event else "",
            a.track or "",
            a.presentation_type.value if a.presentation_type else "",
            a.status.value.replace("_", " ").title() if a.status else "",
            a.word_count or 0,
            a.keywords or "",
            f"{a.submitter.firstname} {a.submitter.lastname}" if a.submitter else "",
            a.submitter.email if a.submitter else "",
            f"{first.firstname} {first.lastname}" if first else "",
            first.email or "" if first else "",
            first.affiliation or "" if first else "",
            first.country or "" if first else "",
            all_authors,
            len(assignments),
            len(reviews),
            avg("relevance_score"),
            avg("methodology_score"),
            avg("originality_score"),
            avg("overall_score"),
            a.created_at.strftime("%d %b %Y") if a.created_at else "",
        ]

        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            cell.font = body_font
            cell.fill = use_fill
            cell.alignment = left

    # Auto column widths
    col_widths = [6, 40, 30, 20, 16, 16, 10, 30, 22, 28, 22, 28, 28, 16, 50, 10, 10, 12, 12, 12, 12, 14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: Authors ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Authors")
    ws2.sheet_properties.tabColor = "0095B6"
    auth_headers = ["Abstract #", "Abstract Title", "Author Order", "First Name", "Last Name",
                    "Email", "Affiliation", "Country", "Is Presenting"]
    ws2.row_dimensions[1].height = 22
    for ci, h in enumerate(auth_headers, 1):
        cell = ws2.cell(1, ci, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws2.freeze_panes = "A2"
    ri2 = 2
    for a in abstracts:
        for au in sorted(a.authors, key=lambda x: x.author_order):
            use_fill = alt_fill if ri2 % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
            row = [a.id, a.title, au.author_order + 1, au.firstname, au.lastname,
                   au.email or "", au.affiliation or "", au.country or "",
                   "Yes" if au.is_presenting else "No"]
            for ci, val in enumerate(row, 1):
                cell = ws2.cell(ri2, ci, val)
                cell.font = body_font
                cell.fill = use_fill
                cell.alignment = left
            ri2 += 1
    for ci, w in enumerate([10, 40, 12, 16, 16, 28, 28, 16, 12], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 3: Reviews ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Reviews")
    ws3.sheet_properties.tabColor = "0095B6"
    rev_headers = ["Abstract #", "Abstract Title", "Reviewer Name", "Reviewer Email",
                   "Completed", "Relevance", "Methodology", "Originality",
                   "Overall", "Recommendation", "Comments", "Review Date"]
    ws3.row_dimensions[1].height = 22
    for ci, h in enumerate(rev_headers, 1):
        cell = ws3.cell(1, ci, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws3.freeze_panes = "A2"
    ri3 = 2
    for a in abstracts:
        for ra in a.reviewer_assignments:
            use_fill = alt_fill if ri3 % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
            rv = ra.review
            row = [
                a.id, a.title,
                f"{ra.reviewer.firstname} {ra.reviewer.lastname}" if ra.reviewer else "",
                ra.reviewer.email if ra.reviewer else "",
                "Yes" if ra.completed else "No",
                rv.relevance_score if rv else "",
                rv.methodology_score if rv else "",
                rv.originality_score if rv else "",
                rv.overall_score if rv else "",
                rv.recommendation.value.replace("_", " ").title() if rv else "",
                rv.comments if rv else "",
                rv.submitted_at.strftime("%d %b %Y") if rv and rv.submitted_at else "",
            ]
            for ci, val in enumerate(row, 1):
                cell = ws3.cell(ri3, ci, val)
                cell.font = body_font
                cell.fill = use_fill
                cell.alignment = left
            ri3 += 1
    for ci, w in enumerate([10, 40, 22, 28, 10, 10, 12, 12, 10, 18, 50, 12], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    # Style sheet tabs
    ws.sheet_properties.tabColor = "0095B6"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "abstracts_export.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/my-submissions")
@router.get("/my-submissions/")
def my_submissions(
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    abstracts = db.query(Abstract).options(
        joinedload(Abstract.authors),
        joinedload(Abstract.event),
        joinedload(Abstract.reviewer_assignments),
    ).filter(
        Abstract.submitted_by == current_user["user_id"],
        Abstract.deleted_at == None,
    ).order_by(Abstract.created_at.desc()).all()
    return [_serialize_abstract(a) for a in abstracts]


@router.get("/my-presenter-status")
def my_presenter_status(
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    """
    Returns whether the current user is an accepted abstract presenter
    and whether they have a paid event registration.
    Used to gate access to presentation templates.
    """
    uid = current_user["user_id"]

    # Check: has any accepted abstract where user is the presenting author
    is_presenting = (
        db.query(AbstractAuthor)
        .join(Abstract, AbstractAuthor.abstract_id == Abstract.id)
        .join(User, User.email == AbstractAuthor.email)
        .filter(
            User.id == uid,
            Abstract.status == "accepted",
            Abstract.deleted_at == None,
            AbstractAuthor.is_presenting == True,
        )
        .first()
        is not None
    )

    # Check: has any paid registration for any event
    has_paid = (
        db.query(Registration)
        .filter(Registration.user_id == uid, Registration.paid == True)
        .first()
        is not None
    )

    return {
        "is_presenting": is_presenting,
        "has_paid": has_paid,
        "is_paid_presenter": is_presenting and has_paid,
    }


@router.get("/my-reviews")
def my_reviews(
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    assignments = db.query(AbstractReviewer).options(
        joinedload(AbstractReviewer.abstract).joinedload(Abstract.authors),
        joinedload(AbstractReviewer.abstract).joinedload(Abstract.event),
        joinedload(AbstractReviewer.review),
    ).filter(
        AbstractReviewer.reviewer_id == current_user["user_id"],
    ).all()
    return [
        {
            "assignment_id": a.id,
            "assigned_at": a.assigned_at,
            "completed": a.completed,
            "abstract": _serialize_abstract(a.abstract),
            "review": {
                "id": a.review.id,
                "relevance_score": a.review.relevance_score,
                "methodology_score": a.review.methodology_score,
                "originality_score": a.review.originality_score,
                "overall_score": a.review.overall_score,
                "recommendation": a.review.recommendation.value,
                "comments": a.review.comments,
                "submitted_at": a.review.submitted_at,
            } if a.review else None,
        }
        for a in assignments
    ]


@router.get("/reviewers")
def list_reviewers(
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
    event_id: int = None,
):
    auth_dependency.secure_access("MANAGE_REVIEWERS", current_user["user_id"])
    q = db.query(AbstractReviewer).options(
        joinedload(AbstractReviewer.reviewer),
        joinedload(AbstractReviewer.abstract).joinedload(Abstract.event),
        joinedload(AbstractReviewer.review),
    )
    if event_id:
        from sqlalchemy import join
        q = q.join(Abstract, AbstractReviewer.abstract_id == Abstract.id).filter(Abstract.event_id == event_id)
    assignments = q.all()

    reviewers: dict = {}
    for a in assignments:
        rid = a.reviewer_id
        if rid not in reviewers:
            reviewers[rid] = {
                "reviewer_id": rid,
                "name": f"{a.reviewer.firstname} {a.reviewer.lastname}" if a.reviewer else "",
                "email": a.reviewer.email if a.reviewer else "",
                "total": 0,
                "completed": 0,
                "assignments": [],
            }
        reviewers[rid]["total"] += 1
        if a.completed:
            reviewers[rid]["completed"] += 1
        reviewers[rid]["assignments"].append({
            "assignment_id": a.id,
            "abstract_id": a.abstract_id,
            "abstract_title": a.abstract.title if a.abstract else "",
            "event": a.abstract.event.event if a.abstract and a.abstract.event else "",
            "event_id": a.abstract.event_id if a.abstract else None,
            "assigned_at": a.assigned_at,
            "completed": a.completed,
        })
    return list(reviewers.values())


@router.get("/presenter-registration-status")
def presenter_registration_status(
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
    event_id: int = Query(None),
):
    """
    Returns registration + payment status keyed by presenter email.
    Used to show badges and enable filtering on the abstracts list.
    """
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])

    q = (
        db.query(AbstractAuthor)
        .join(Abstract, AbstractAuthor.abstract_id == Abstract.id)
        .options(joinedload(AbstractAuthor.abstract))
        .filter(
            AbstractAuthor.is_presenting == True,
            AbstractAuthor.email != None,
            AbstractAuthor.email != "",
            Abstract.status == "accepted",
            Abstract.deleted_at == None,
        )
    )
    if event_id:
        q = q.filter(Abstract.event_id == event_id)

    status_map = {}
    for pa in q.all():
        email = pa.email.strip().lower()
        if email in status_map:
            continue
        user = db.query(User).filter(User.email == email).first()
        has_account = user is not None
        has_registered = False
        has_paid = False
        if has_account:
            target_event_id = event_id or pa.abstract.event_id
            reg = db.query(Registration).filter(
                Registration.user_id == user.id,
                Registration.event_id == target_event_id,
            ).first()
            if reg:
                has_registered = True
                has_paid = bool(reg.paid)
        status_map[email] = {
            "has_account": has_account,
            "has_registered": has_registered,
            "has_paid": has_paid,
        }
    return status_map


@router.get("/registration-reminder-preview")
def registration_reminder_preview(
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
    event_id: int = Query(None),
):
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])

    q = db.query(AbstractAuthor).join(
        Abstract, AbstractAuthor.abstract_id == Abstract.id
    ).options(
        joinedload(AbstractAuthor.abstract).joinedload(Abstract.event),
    ).filter(
        AbstractAuthor.is_presenting == True,
        AbstractAuthor.email != None,
        AbstractAuthor.email != "",
        Abstract.status == "accepted",
        Abstract.deleted_at == None,
    )
    if event_id:
        q = q.filter(Abstract.event_id == event_id)

    presenters = q.all()

    # Deduplicate by email (a presenter may have multiple abstracts)
    seen_emails = set()
    results = []
    for pa in presenters:
        email = pa.email.strip().lower()
        if email in seen_emails:
            continue
        seen_emails.add(email)

        user = db.query(User).filter(User.email == email).first()
        has_account = user is not None
        has_registration = False
        if has_account:
            has_registration = db.query(Registration).filter(
                Registration.user_id == user.id,
                Registration.event_id == (event_id or pa.abstract.event_id),
            ).first() is not None

        results.append({
            "firstname": pa.firstname,
            "lastname": pa.lastname,
            "email": pa.email,
            "abstract_title": pa.abstract.title,
            "abstract_id": pa.abstract_id,
            "has_account": has_account,
            "has_registered": has_registration,
        })

    # Only return those who haven't registered
    unregistered = [r for r in results if not r["has_registered"]]
    return unregistered


@router.post("/send-registration-reminders")
def send_registration_reminders(
    current_user: user_dependency,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
    event_id: int = Query(None),
):
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])

    q = db.query(AbstractAuthor).join(
        Abstract, AbstractAuthor.abstract_id == Abstract.id
    ).options(
        joinedload(AbstractAuthor.abstract).joinedload(Abstract.event),
    ).filter(
        AbstractAuthor.is_presenting == True,
        AbstractAuthor.email != None,
        AbstractAuthor.email != "",
        Abstract.status == "accepted",
        Abstract.deleted_at == None,
    )
    if event_id:
        q = q.filter(Abstract.event_id == event_id)

    presenters = q.all()

    seen_emails = set()
    recipients = []
    for pa in presenters:
        email = pa.email.strip().lower()
        if email in seen_emails:
            continue
        seen_emails.add(email)

        user = db.query(User).filter(User.email == email).first()
        if user:
            has_registration = db.query(Registration).filter(
                Registration.user_id == user.id,
                Registration.event_id == (event_id or pa.abstract.event_id),
            ).first() is not None
            if has_registration:
                continue

        recipients.append({
            "firstname": pa.firstname,
            "lastname": pa.lastname,
            "email": email,
            "abstract_title": pa.abstract.title,
            "has_account": user is not None,
        })

    event = db.query(Event).filter(Event.id == event_id).first() if event_id else None
    event_name = event.event if event else "the conference"

    import utils.mailer_util as mailer_util
    for r in recipients:
        mailer_util.registration_reminder_email(
            recipient_email=r["email"],
            firstname=r["firstname"],
            lastname=r["lastname"],
            event_name=event_name,
            has_account=r["has_account"],
            background_tasks=background_tasks,
        )

    return {
        "sent": len(recipients),
        "message": f"Registration reminders queued for {len(recipients)} presenter(s).",
    }


@router.post("/import-preview")
async def import_abstracts_preview(
    file: UploadFile = File(...),
    event_id: int = Query(None),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
):
    """
    Dry-run parse an ODS/XLSX file and return what would be imported
    without writing to the database.
    """
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])

    rows, parse_error = await _parse_import_file(file)
    if parse_error:
        raise HTTPException(status_code=400, detail=parse_error)

    # Resolve event
    event = _resolve_event(db, event_id, rows)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found. Pass event_id or ensure an event name matches.")

    # Get existing abstract titles for duplicate detection
    existing_titles = {
        a.title.strip().lower()
        for a in db.query(Abstract.title).filter(Abstract.event_id == event.id, Abstract.deleted_at == None).all()
    }

    preview = []
    for row in rows:
        title = str(row.get("Title") or "").strip()
        if not title:
            continue
        presenter_email = str(row.get("Presenter Email") or "").strip().lower()
        user = db.query(User).filter(User.email == presenter_email).first() if presenter_email else None
        preview.append({
            "title": title,
            "presenter_name": str(row.get("Presenter Name") or "").strip(),
            "presenter_email": presenter_email,
            "presentation_type": _map_ptype(str(row.get("Status") or "")),
            "track": str(row.get("Topic") or "").strip(),
            "keywords": str(row.get("Keywords") or "").strip(),
            "is_duplicate": title.lower() in existing_titles,
            "has_account": user is not None,
            "user_id": user.id if user else None,
            "author_count": _count_authors(row),
        })

    total = len(preview)
    to_import = sum(1 for r in preview if not r["is_duplicate"])
    duplicates = total - to_import
    with_accounts = sum(1 for r in preview if r["has_account"])

    return {
        "event_id": event.id,
        "event_name": event.event,
        "total": total,
        "to_import": to_import,
        "duplicates": duplicates,
        "with_accounts": with_accounts,
        "without_accounts": total - with_accounts,
        "rows": preview,
    }


@router.post("/import")
async def import_abstracts(
    file: UploadFile = File(...),
    event_id: int = Query(None),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
):
    """
    Import abstracts from an ODS/XLSX file into the database.
    Skips duplicates (same title + event). Marks presenting author with is_presenting=True.
    Falls back to the current admin user as submitted_by when presenter has no account.
    """
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])

    rows, parse_error = await _parse_import_file(file)
    if parse_error:
        raise HTTPException(status_code=400, detail=parse_error)

    event = _resolve_event(db, event_id, rows)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found.")

    existing_titles = {
        a.title.strip().lower()
        for a in db.query(Abstract.title).filter(Abstract.event_id == event.id, Abstract.deleted_at == None).all()
    }

    imported = 0
    skipped_duplicate = 0
    errors = []

    for idx, row in enumerate(rows):
        title = str(row.get("Title") or "").strip()
        if not title:
            continue
        if title.lower() in existing_titles:
            skipped_duplicate += 1
            continue

        try:
            presenter_email = str(row.get("Presenter Email") or "").strip().lower()
            presenter_name = str(row.get("Presenter Name") or "").strip()
            abstract_text = str(row.get("Description") or "").strip()
            keywords = str(row.get("Keywords") or "").strip()
            track = str(row.get("Topic") or "").strip()
            ptype = _map_ptype(str(row.get("Status") or ""))

            # Find presenter user for submitted_by; fall back to current admin
            submitter = db.query(User).filter(User.email == presenter_email).first() if presenter_email else None
            submitted_by = submitter.id if submitter else current_user["user_id"]

            abstract = Abstract(
                event_id=event.id,
                submitted_by=submitted_by,
                title=title,
                abstract_text=abstract_text or "(imported)",
                keywords=keywords,
                track=track,
                presentation_type=ptype,
                status="accepted",
                word_count=len(abstract_text.split()) if abstract_text else 0,
            )
            db.add(abstract)
            db.flush()

            # Parse authors (pipe-separated in Author Name / Author Affiliation)
            author_names = [n.strip() for n in str(row.get("Author Name") or "").split("|") if n.strip()]
            affiliations = [a.strip() for a in str(row.get("Author Affiliation") or "").split("|") if True]
            # Pad affiliations list to match author count
            while len(affiliations) < len(author_names):
                affiliations.append("")

            # Regex to strip honorific prefixes (with or without trailing period)
            _title_prefix = re.compile(
                r'^(Dr\.?|Prof\.?|Mr\.?|Mrs\.?|Ms\.?|Assoc\.?|Prof\.?)\s+',
                flags=re.IGNORECASE,
            )

            def _clean_name(raw: str):
                """Return (firstname, lastname) cleaned from a raw author name."""
                name = re.sub(r'\d+$', '', raw).strip()
                parts = name.split(None, 1)
                fn = _title_prefix.sub('', parts[0] if parts else name).strip()[:99]
                ln = (parts[1] if len(parts) > 1 else "")[:99]
                # If lastname is suspiciously long it likely contains all authors → use full name
                if len(ln) > 60:
                    fn = name[:99]
                    ln = ""
                return fn, ln

            def _valid_email(val):
                """Return val if it looks like an email, else None."""
                v = str(val or "").strip()
                return v if "@" in v and "." in v.split("@")[-1] else None

            presenter_marked = False
            for order, name in enumerate(author_names):
                firstname, lastname = _clean_name(name)

                # Check if this author is the presenter
                is_presenting = (
                    presenter_name.lower() in name.lower() or
                    name.lower() in presenter_name.lower()
                )
                if is_presenting:
                    presenter_marked = True

                # Get email: for presenting author, prefer Presenter Email over Author Email column
                email_col = f"Author Email_{order + 1}"
                raw_author_email = _valid_email(row.get(email_col))
                if is_presenting:
                    # Presenter Email is the authoritative source; fall back to column value
                    email = _valid_email(presenter_email) or raw_author_email
                else:
                    email = raw_author_email

                db.add(AbstractAuthor(
                    abstract_id=abstract.id,
                    firstname=firstname,
                    lastname=lastname,
                    email=email,
                    affiliation=(affiliations[order] if order < len(affiliations) else "")[:299],
                    country=None,
                    is_presenting=is_presenting,
                    author_order=order,
                ))

            # If no author was matched as presenter, add presenter as a standalone record
            if not presenter_marked:
                firstname, lastname = _clean_name(presenter_name)
                db.add(AbstractAuthor(
                    abstract_id=abstract.id,
                    firstname=firstname,
                    lastname=lastname,
                    email=_valid_email(presenter_email),
                    affiliation="",
                    country=None,
                    is_presenting=True,
                    author_order=len(author_names),
                ))

            existing_titles.add(title.lower())
            imported += 1

        except Exception as e:
            errors.append({"row": idx + 2, "title": title[:60], "error": str(e)})

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database commit failed: {str(e)}")

    return {
        "success": True,
        "event_id": event.id,
        "event_name": event.event,
        "imported": imported,
        "skipped_duplicate": skipped_duplicate,
        "errors": errors,
        "message": f"Imported {imported} abstract(s). Skipped {skipped_duplicate} duplicate(s)."
            + (f" {len(errors)} error(s) — check 'errors' field." if errors else ""),
    }


# ── Helpers ──────────────────────────────────────────────────────────────────

async def _parse_import_file(file: UploadFile):
    """Read uploaded ODS or XLSX and return (list[dict], error_string|None)."""
    import pandas as pd

    content = await file.read()
    fname = (file.filename or "").lower()
    try:
        if fname.endswith(".ods") or "ods" in fname:
            df = pd.read_excel(io.BytesIO(content), engine="odf")
            # Fix mojibake: some ODS files have UTF-8 bytes mis-stored as latin-1 strings.
            # Try re-encoding latin-1 → UTF-8; silently skip values that are already correct.
            def _fix_encoding(v):
                if not isinstance(v, str):
                    return v
                try:
                    return v.encode("latin-1").decode("utf-8")
                except (UnicodeEncodeError, UnicodeDecodeError):
                    return v
            for col in df.select_dtypes(include="object").columns:
                df[col] = df[col].apply(_fix_encoding)
        else:
            df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        return [], f"Could not parse file: {e}"

    # Drop fully-empty rows and the unnamed trailing columns
    df = df.dropna(how="all")
    cols_to_keep = [c for c in df.columns if not str(c).startswith("Unnamed")]
    df = df[cols_to_keep]

    return df.to_dict("records"), None


def _resolve_event(db: Session, event_id, rows: list):
    """Find event by id, or by matching the event name from the first data row."""
    if event_id:
        return db.query(Event).filter(Event.id == event_id, Event.deleted_at == None).first()
    # Try to match by name from the data
    if rows:
        event_name = str(rows[0].get("Event Name") or "").strip()
        if event_name:
            ev = db.query(Event).filter(
                Event.event.ilike(f"%{event_name[:40]}%"),
                Event.deleted_at == None,
            ).first()
            if ev:
                return ev
    # Fall back to the first active event
    return db.query(Event).filter(Event.deleted_at == None).order_by(Event.id).first()


def _map_ptype(status_str: str) -> str:
    sl = status_str.lower()
    if "poster" in sl:
        return "poster"
    return "oral"


def _count_authors(row: dict) -> int:
    names = str(row.get("Author Name") or "")
    return max(1, len([n for n in names.split("|") if n.strip()]))


@router.post("/send-registration-reminders")
def send_registration_reminders(
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    body: dict = None,
):
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])

    event_id = (body or {}).get("event_id")

    q = db.query(AbstractAuthor).join(
        Abstract, AbstractAuthor.abstract_id == Abstract.id
    ).options(
        joinedload(AbstractAuthor.abstract).joinedload(Abstract.event),
    ).filter(
        AbstractAuthor.is_presenting == True,
        AbstractAuthor.email != None,
        AbstractAuthor.email != "",
        Abstract.status == "accepted",
        Abstract.deleted_at == None,
    )
    if event_id:
        q = q.filter(Abstract.event_id == event_id)

    presenters = q.all()

    seen_emails = set()
    reminders_sent = 0
    skipped_registered = 0

    for pa in presenters:
        email = pa.email.strip().lower()
        if email in seen_emails:
            continue
        seen_emails.add(email)

        user = db.query(User).filter(User.email == email).first()
        has_account = user is not None

        target_event_id = event_id or pa.abstract.event_id
        has_registration = False
        if has_account:
            has_registration = db.query(Registration).filter(
                Registration.user_id == user.id,
                Registration.event_id == target_event_id,
            ).first() is not None

        if has_registration:
            skipped_registered += 1
            continue

        # Send the reminder email
        event_name = pa.abstract.event.event if pa.abstract and pa.abstract.event else "ECSACONM Event"
        firstname = pa.firstname or "Presenter"

        import utils.mailer_util as mailer_util
        from models.models import EmailTemplate as EmailTemplateModel
        from jinja2 import Template as Jinja2Template

        if has_account:
            subject = f"Reminder: Register for {event_name}"
        else:
            subject = f"Action Required: Create Account & Register for {event_name}"

        # Load template from DB first, fall back to file
        db_tpl = db.query(EmailTemplateModel).filter_by(template_key="registration_reminder").first()
        render_vars = dict(
            subject=subject,
            firstname=firstname,
            event_name=event_name,
            abstract_title=pa.abstract.title,
            has_account=has_account,
            year=mailer_util.YEAR,
        )
        try:
            if db_tpl and db_tpl.body_html:
                email_body = Jinja2Template(db_tpl.body_html).render(**render_vars)
            else:
                file_tpl = mailer_util.templates.get_template("registration_reminder_template.html")
                email_body = file_tpl.render(**render_vars)
        except Exception:
            email_body = (
                f"<p>Dear {firstname},</p>"
                f"<p>Your abstract has been accepted for <strong>{event_name}</strong>. "
                f"Please complete your registration.</p>"
                f"<p>ECSACONM Events Team</p>"
            )

        mailer_util.send_email_backgroundable(email, subject, email_body, background_tasks)
        reminders_sent += 1

    return {
        "message": f"Reminders sent: {reminders_sent}, skipped (already registered): {skipped_registered}",
        "reminders_sent": reminders_sent,
        "skipped_registered": skipped_registered,
    }


@router.get("/{abstract_id}")
def get_abstract(
    abstract_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    abstract = db.query(Abstract).options(
        joinedload(Abstract.authors),
        joinedload(Abstract.submitter),
        joinedload(Abstract.event),
        joinedload(Abstract.reviewer_assignments).joinedload(AbstractReviewer.reviewer),
        joinedload(Abstract.reviewer_assignments).joinedload(AbstractReviewer.review),
    ).filter(Abstract.id == abstract_id, Abstract.deleted_at == None).first()
    if not abstract:
        raise HTTPException(status_code=404, detail="Abstract not found")
    return _serialize_abstract(abstract)


@router.put("/{abstract_id}")
def update_abstract(
    abstract_id: int,
    schema: AbstractUpdateSchema,
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    abstract = db.query(Abstract).filter(Abstract.id == abstract_id, Abstract.deleted_at == None).first()
    if not abstract:
        raise HTTPException(status_code=404, detail="Abstract not found")

    if schema.title: abstract.title = schema.title
    if schema.abstract_text:
        abstract.abstract_text = schema.abstract_text
        abstract.word_count = len(schema.abstract_text.split())
    if schema.keywords is not None: abstract.keywords = schema.keywords
    if schema.track is not None: abstract.track = schema.track
    if schema.presentation_type: abstract.presentation_type = schema.presentation_type
    if schema.status: abstract.status = schema.status

    if schema.authors is not None:
        db.query(AbstractAuthor).filter(AbstractAuthor.abstract_id == abstract_id).delete()
        for i, author in enumerate(schema.authors):
            db.add(AbstractAuthor(
                abstract_id=abstract_id,
                firstname=author.firstname,
                lastname=author.lastname,
                email=author.email,
                affiliation=author.affiliation,
                country=author.country,
                is_presenting=author.is_presenting,
                author_order=i,
            ))

    db.commit()
    db.refresh(abstract)
    abstract = db.query(Abstract).options(
        joinedload(Abstract.authors),
        joinedload(Abstract.submitter),
        joinedload(Abstract.event),
        joinedload(Abstract.reviewer_assignments).joinedload(AbstractReviewer.reviewer),
        joinedload(Abstract.reviewer_assignments).joinedload(AbstractReviewer.review),
    ).filter(Abstract.id == abstract_id).first()
    return _serialize_abstract(abstract)


@router.put("/{abstract_id}/status")
def update_status(
    abstract_id: int,
    payload: dict,
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    abstract = db.query(Abstract).filter(Abstract.id == abstract_id, Abstract.deleted_at == None).first()
    if not abstract:
        raise HTTPException(status_code=404, detail="Abstract not found")
    abstract.status = payload.get("status", abstract.status)
    db.commit()
    return {"message": "Status updated"}


@router.post("/{abstract_id}/assign-reviewer", status_code=status.HTTP_201_CREATED)
def assign_reviewer(
    abstract_id: int,
    schema: AssignReviewerSchema,
    background_tasks: BackgroundTasks,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
):
    auth_dependency.secure_access("MANAGE_REVIEWERS", current_user["user_id"])
    abstract = db.query(Abstract).options(
        joinedload(Abstract.event)
    ).filter(Abstract.id == abstract_id, Abstract.deleted_at == None).first()
    if not abstract:
        raise HTTPException(status_code=404, detail="Abstract not found")

    existing = db.query(AbstractReviewer).filter(
        AbstractReviewer.abstract_id == abstract_id,
        AbstractReviewer.reviewer_id == schema.reviewer_id,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Reviewer already assigned")

    reviewer = db.query(User).filter(User.id == schema.reviewer_id).first()
    if not reviewer:
        raise HTTPException(status_code=404, detail="Reviewer not found")

    assignment = AbstractReviewer(
        abstract_id=abstract_id,
        reviewer_id=schema.reviewer_id,
        assigned_by=current_user["user_id"],
    )
    db.add(assignment)

    # Generate fresh credentials and email them with the abstract context
    new_password = auth_dependency.generate_random_password()
    reviewer.hashed_password = auth_dependency.hash_password(new_password)
    db.commit()

    import utils.mailer_util as mailer_util
    mailer_util.reviewer_assignment_email(
        recipient_email=reviewer.email,
        firstname=reviewer.firstname,
        password=new_password,
        abstract_title=abstract.title,
        event_name=abstract.event.event if abstract.event else None,
        background_tasks=background_tasks,
    )

    return {"message": "Reviewer assigned", "reviewer_name": f"{reviewer.firstname} {reviewer.lastname}"}


@router.delete("/{abstract_id}/reviewers/{reviewer_id}")
def remove_reviewer(
    abstract_id: int,
    reviewer_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
):
    auth_dependency.secure_access("MANAGE_REVIEWERS", current_user["user_id"])
    assignment = db.query(AbstractReviewer).filter(
        AbstractReviewer.abstract_id == abstract_id,
        AbstractReviewer.reviewer_id == reviewer_id,
    ).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    db.delete(assignment)
    db.commit()
    return {"message": "Reviewer removed"}


@router.delete("/{abstract_id}")
def delete_abstract(
    abstract_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
):
    auth_dependency.secure_access("MANAGE_REVIEWERS", current_user["user_id"])
    abstract = db.query(Abstract).filter(Abstract.id == abstract_id, Abstract.deleted_at == None).first()
    if not abstract:
        raise HTTPException(status_code=404, detail="Abstract not found")
    abstract.deleted_at = datetime.utcnow()
    db.commit()
    return {"message": "Abstract deleted"}


@router.post("/reviewers/create", status_code=status.HTTP_201_CREATED)
def create_reviewer(
    schema: CreateReviewerSchema,
    current_user: user_dependency,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
):
    """Create a new user account designated as a reviewer and send them login credentials."""
    auth_dependency.secure_access("MANAGE_REVIEWERS", current_user["user_id"])

    if db.query(User).filter(User.email == schema.email).first():
        raise HTTPException(status_code=400, detail="A user with this email already exists")

    # Get or create the phone (phone optional for reviewers — use email as fallback stub)
    phone = schema.phone or schema.email

    password = auth_dependency.generate_random_password()
    hashed = auth_dependency.hash_password(password)

    user = User(
        firstname=schema.firstname,
        lastname=schema.lastname,
        email=schema.email,
        phone=phone,
        hashed_password=hashed,
        verified=True,
    )
    db.add(user)
    db.flush()

    # Assign "User" role so they can log in
    user_role = db.query(Role).filter(Role.role == "User").first()
    if user_role:
        db.add(UserRole(user_id=user.id, role_id=user_role.id))

    db.commit()
    db.refresh(user)

    # Credentials are intentionally NOT sent here — they will be emailed
    # automatically when the reviewer is first assigned to an abstract.

    return {
        "id": user.id,
        "firstname": user.firstname,
        "lastname": user.lastname,
        "email": user.email,
    }


@router.post("/reviews/{assignment_id}", status_code=status.HTTP_201_CREATED)
def submit_review(
    assignment_id: int,
    schema: AbstractReviewSchema,
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    assignment = db.query(AbstractReviewer).filter(
        AbstractReviewer.id == assignment_id,
        AbstractReviewer.reviewer_id == current_user["user_id"],
    ).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    if assignment.review:
        raise HTTPException(status_code=400, detail="Review already submitted")

    review = AbstractReview(
        assignment_id=assignment_id,
        relevance_score=schema.relevance_score,
        methodology_score=schema.methodology_score,
        originality_score=schema.originality_score,
        overall_score=schema.overall_score,
        recommendation=schema.recommendation,
        comments=schema.comments,
        confidential_comments=schema.confidential_comments,
    )
    db.add(review)
    assignment.completed = True
    db.commit()
    return {"message": "Review submitted successfully"}


@router.get("/{abstract_id}/reviews")
def get_reviews(
    abstract_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    assignments = db.query(AbstractReviewer).options(
        joinedload(AbstractReviewer.reviewer),
        joinedload(AbstractReviewer.review),
    ).filter(AbstractReviewer.abstract_id == abstract_id).all()
    return [
        {
            "assignment_id": a.id,
            "reviewer": f"{a.reviewer.firstname} {a.reviewer.lastname}" if a.reviewer else None,
            "completed": a.completed,
            "review": {
                "relevance_score": a.review.relevance_score,
                "methodology_score": a.review.methodology_score,
                "originality_score": a.review.originality_score,
                "overall_score": a.review.overall_score,
                "recommendation": a.review.recommendation.value,
                "comments": a.review.comments,
                "submitted_at": a.review.submitted_at,
            } if a.review else None,
        }
        for a in assignments
    ]


# ── Presenter: upload presentation file ──────────────────────────────────────

@router.post("/{abstract_id}/upload-presentation", status_code=status.HTTP_200_OK)
async def upload_presentation(
    abstract_id: int,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Presenter uploads their presentation file for a given abstract."""
    abstract = (
        db.query(Abstract)
        .filter(Abstract.id == abstract_id, Abstract.deleted_at == None)
        .first()
    )
    if not abstract:
        raise HTTPException(status_code=404, detail="Abstract not found")

    # Only the submitter (or an admin) may upload
    if abstract.submitted_by != current_user["user_id"]:
        # allow admin to upload on behalf of presenter
        from models.models import UserRole, Role
        is_admin = (
            db.query(UserRole)
            .join(Role, UserRole.role_id == Role.id)
            .filter(UserRole.user_id == current_user["user_id"], Role.role == "Admin")
            .first()
        )
        if not is_admin:
            raise HTTPException(status_code=403, detail="Not authorised to upload for this abstract")

    ext = os.path.splitext(file.filename or "")[-1].lower()
    if ext not in ALLOWED_PRESENTATION_EXTS:
        raise HTTPException(
            status_code=400,
            detail=f"Allowed types: {', '.join(ALLOWED_PRESENTATION_EXTS)}",
        )

    content = await file.read()
    if len(content) > MAX_PRESENTATION_MB * 1024 * 1024:
        raise HTTPException(status_code=400, detail=f"File must be under {MAX_PRESENTATION_MB} MB")

    # Remove previous file if it exists
    if abstract.presentation_file and os.path.exists(abstract.presentation_file):
        try:
            os.remove(abstract.presentation_file)
        except OSError:
            pass

    stored_name = f"{uuid.uuid4().hex}{ext}"
    stored_path = os.path.join(PRESENTATION_UPLOAD_DIR, stored_name)
    with open(stored_path, "wb") as fh:
        fh.write(content)

    abstract.presentation_file = stored_path
    abstract.presentation_uploaded_at = datetime.now(timezone.utc)
    db.commit()

    return {
        "message": "Presentation uploaded successfully",
        "filename": file.filename,
        "abstract_id": abstract_id,
    }


# ── Admin: list all uploaded presentations ────────────────────────────────────

@router.get("/uploaded-presentations/list")
def list_uploaded_presentations(
    current_user: user_dependency,
    skip: int = 0,
    limit: int = 50,
    search: str = "",
    db: Session = Depends(get_db),
):
    """Returns all abstracts that have a presentation file uploaded."""
    q = (
        db.query(Abstract)
        .options(
            joinedload(Abstract.event),
            joinedload(Abstract.submitter),
            joinedload(Abstract.authors),
        )
        .filter(
            Abstract.deleted_at == None,
            Abstract.presentation_file != None,
        )
    )
    if search:
        q = q.filter(Abstract.title.ilike(f"%{search}%"))

    total = q.count()
    rows = q.order_by(Abstract.presentation_uploaded_at.desc()).offset(skip).limit(limit).all()

    return {
        "total": total,
        "data": [
            {
                "id": a.id,
                "title": a.title,
                "status": a.status.value if a.status else None,
                "event": a.event.event if a.event else None,
                "submitter_name": f"{a.submitter.firstname} {a.submitter.lastname}" if a.submitter else None,
                "submitter_email": a.submitter.email if a.submitter else None,
                "presentation_file": a.presentation_file,
                "presentation_uploaded_at": (
                    a.presentation_uploaded_at.isoformat()
                    if a.presentation_uploaded_at else None
                ),
                "presenting_author": next(
                    (
                        {"name": f"{au.firstname} {au.lastname}", "email": au.email}
                        for au in a.authors if au.is_presenting
                    ),
                    None,
                ),
            }
            for a in rows
        ],
    }


# ── Admin: download a presenter's uploaded file ───────────────────────────────

@router.get("/{abstract_id}/download-presentation")
def download_presentation(
    abstract_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    abstract = (
        db.query(Abstract)
        .filter(Abstract.id == abstract_id, Abstract.deleted_at == None)
        .first()
    )
    if not abstract or not abstract.presentation_file:
        raise HTTPException(status_code=404, detail="No presentation file found")
    if not os.path.exists(abstract.presentation_file):
        raise HTTPException(status_code=404, detail="File not found on server")

    ext = os.path.splitext(abstract.presentation_file)[-1]
    download_name = f"presentation_{abstract_id}{ext}"
    return FileResponse(
        path=abstract.presentation_file,
        filename=download_name,
        media_type="application/octet-stream",
    )
