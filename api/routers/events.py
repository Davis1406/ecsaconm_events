import math, os
import uuid
import shutil
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from io import BytesIO
from typing import Literal
from fastapi.responses import JSONResponse
from sqlalchemy import or_
from fastapi import status, HTTPException, File, Form, UploadFile
from typing import Annotated
from core.database import get_db
from sqlalchemy.orm import Session, joinedload
from datetime import datetime
from dependencies.auth_dependency import Auth
from dependencies.dependency import Dependency
from dependencies.auth_dependency import get_current_user, get_optional_current_user
from typing import Optional
from fastapi import APIRouter, HTTPException, Depends, Query, Request
from models.models import Event, User, Registration, Document, Link, Payment, UserProfile, Country, UserPhoto, UserRole
from models.models import ParticipationRole, PaymentStatus, PaymentMethod
from utils.mailer_util import hash_password
from schemas.events_space import EventSchema, EventUpdateSchema, RegistrationSchema, LinkSchema, SendReceiptSchema
from utils.receipt_generator import generate_receipt_pdf
from utils.mailer_util import send_email_with_attachment
from fastapi import BackgroundTasks
from PIL import Image
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
import qrcode
from fastapi.responses import StreamingResponse
from reportlab.pdfgen import canvas
import unicodedata
import re
import urllib.parse


router = APIRouter()

user_dependency = Annotated[dict, Depends(get_current_user)]


def get_dependency(db: Session = Depends(get_db)) -> Dependency:
    return Dependency(db)


def get_auth_dependency(db: Session = Depends(get_db)) -> Auth:
    return Auth(db)


CLIENT_ORIGIN = os.getenv("CLIENT_ORIGIN", "unknown_origin")


EVENT_DOCUMENT_DIR = "uploads/event/documents"
if not os.path.exists(EVENT_DOCUMENT_DIR):
    os.makedirs(EVENT_DOCUMENT_DIR)

EVENT_BANNER_DIR = "uploads/event/banners"
if not os.path.exists(EVENT_BANNER_DIR):
    os.makedirs(EVENT_BANNER_DIR)

ORG_UNIT_LOGO_DIR = "uploads/org_unit/logos"
if not os.path.exists(ORG_UNIT_LOGO_DIR):
    os.makedirs(ORG_UNIT_LOGO_DIR)


def get_object(id: int, db: Session, model):
    data = db.query(model).filter(model.id == id).first()
    if data is None:
        raise HTTPException(
            status_code=404,
            detail=f"{model.__name__} with ID {id} does not exist or has been deleted",
        )
    return data


def sanitize_filename(name: str) -> str:
    # Normalize to remove accents
    name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    # Replace spaces and remove non-word characters
    name = re.sub(r"[^\w\s-]", "", name).strip().replace(" ", "_")
    return name


# Define mapping of participation_role keys to display names
PARTICIPATION_ROLE_MAP = {
    "secretariat": "ECSA-HC secretariat",
    "moh": "Country delegate (from Ministry of Health)",
    "member_state": "Participant from ECSA Member States",
    "other_africa": "Participant from other African countries",
    "world": "Participant from the Rest of the World",
    "student": "Student",
    "exibitor": "Sponsor/Exhibitor",
}


def convert_png_to_rgb(path):
    img = Image.open(path)
    if img.mode in ("RGBA", "LA"):
        background = Image.new("RGB", img.size, (255, 255, 255))
        background.paste(img, mask=img.split()[3])  # Use alpha channel as mask
        return ImageReader(background)
    return ImageReader(img)


def normalize_event_name(name: str) -> str:
    return (
        name.replace("ᵗʰ", "th")
        .replace("ˢᵗ", "st")
        .replace("ⁿᵈ", "nd")
        .replace("ʳᵈ", "rd")
    )


@router.get("/active/")
async def get_active_events(
    db: Session = Depends(get_db),
    skip: int = Query(default=0, ge=0),
    limit: int = 10,
    search: str = "",
):
    """Public endpoint — no auth required. Returns upcoming/active events."""
    from datetime import datetime, timezone
    now = datetime.now(timezone.utc)

    search_filter = or_(
        Event.event.ilike(f"%{search}%"),
        Event.theme.ilike(f"%{search}%"),
        Event.description.ilike(f"%{search}%"),
    )

    filters = [Event.deleted_at.is_(None), Event.end_date >= now]
    if search:
        filters.insert(0, search_filter)

    events_query = (
        db.query(Event)
        .options(joinedload(Event.org_unit))
        .filter(*filters)
        .order_by(Event.start_date.asc())
    )

    total_count = events_query.count()
    events = events_query.offset(skip).limit(limit).all()
    pages = math.ceil(total_count / limit) if total_count else 1

    return {
        "pages": pages,
        "total": total_count,
        "data": [
            {
                "id": e.id,
                "event": e.event,
                "theme": e.theme,
                "description": e.description,
                "start_date": e.start_date,
                "end_date": e.end_date,
                "location": e.location,
                "banner_image": e.banner_image,
                "org_unit_id": e.org_unit_id,
                "country_id": e.country_id,
            }
            for e in events
        ],
    }


@router.get("")
@router.get("/")
async def get_events(
    request: Request,
    db: Session = Depends(get_db),
    skip: int = Query(default=0, ge=0),
    limit: int = 10,
    search: str = "",
    dependency: Dependency = Depends(get_dependency),
):
    client_ip = dependency.request_ip(request)
    dependency.log_activity(1, "VIEW_EVENTS", "None", client_ip, "Get all events")

    search_filter = or_(
        Event.event.ilike(f"%{search}%"),
        Event.theme.ilike(f"%{search}%"),
        Event.description.ilike(f"%{search}%"),
    )

    filters = [Event.deleted_at.is_(None)]
    if search_filter is not None:
        filters.insert(0, search_filter)

    events_query = db.query(Event).options(joinedload(Event.org_unit)).filter(*filters)

    total_count = events_query.count()
    events = events_query.offset(skip).limit(limit).all()

    pages = math.ceil(total_count / limit)
    return {
        "pages": pages,
        "data": [
            {
                "id": e.id,
                "event": e.event,
                "theme": e.theme,
                "description": e.description,
                "start_date": e.start_date,
                "end_date": e.end_date,
                "location": e.location,
                "banner_image": e.banner_image,
                "organizers": e.organizers,
                "org_unit_id": e.org_unit_id,
                "org_unit": {
                    "id": e.org_unit.id,
                    "name": e.org_unit.name,
                    "primary_color": e.org_unit.primary_color or "#0095B6",
                    "secondary_color": e.org_unit.secondary_color or "#F7941D",
                    "logo": e.org_unit.logo,
                } if e.org_unit else None,
                "country_id": e.country_id,
            }
            for e in events
        ],
    }


@router.post("/")
async def add_event(
    request: Request,
    event_schema: EventSchema,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    auth_dependency.secure_access("ADD_EVENT", current_user["user_id"])

    client_ip = dependency.request_ip(request)
    dependency.log_activity(
        current_user["user_id"],
        "ADD_EVENT",
        current_user["username"],
        client_ip,
        event_schema.event,
    )

    create_event_model = Event(
        org_unit_id=event_schema.org_unit_id,
        country_id=event_schema.country_id,
        user_id=current_user["user_id"],
        event=event_schema.event,
        theme=event_schema.theme,
        description=event_schema.description,
        location=event_schema.location,
        start_date=event_schema.start_date,
        end_date=event_schema.end_date,
        banner_image=event_schema.banner_image,
        organizers=event_schema.organizers,
        participation_info=event_schema.participation_info,
        logistics_info=event_schema.logistics_info,
        sponsors_info=event_schema.sponsors_info,
    )

    db.add(create_event_model)
    db.commit()
    db.refresh(create_event_model)
    return {"id": create_event_model.id}


@router.get("/scan/{registration_id}")
async def scan_registration(
    registration_id: int,
    db: Session = Depends(get_db),
):
    """Public — called when a badge QR code is scanned."""
    from datetime import date
    from sqlalchemy import Date as SADate
    from models.models import EventAttendance

    registration = db.query(Registration).filter(Registration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    user = registration.user
    event = db.query(Event).filter(Event.id == registration.event_id).first()
    profile = user.user_profile[0] if user and user.user_profile else None
    country = profile.country.country if profile and profile.country else None

    today = date.today()
    today_attendance = (
        db.query(EventAttendance)
        .filter(
            EventAttendance.registration_id == registration_id,
            EventAttendance.created_at.cast(SADate) == today,
        )
        .first()
    )
    all_attendance = (
        db.query(EventAttendance)
        .filter(EventAttendance.registration_id == registration_id)
        .order_by(EventAttendance.created_at.desc())
        .all()
    )
    role_key = (
        registration.participation_role.name
        if hasattr(registration.participation_role, "name")
        else str(registration.participation_role)
    )
    return {
        "registration": {"id": registration.id, "paid": registration.paid, "participation_role": role_key},
        "participant": {
            "firstname": user.firstname if user else "",
            "lastname": user.lastname if user else "",
            "email": user.email if user else "",
            "title": profile.title if profile else "",
            "designation": profile.designation if profile else "",
            "organisation": profile.organisation if profile else "",
            "country": country,
        },
        "event": {
            "id": event.id if event else None,
            "event": event.event if event else "",
            "location": event.location if event else "",
            "start_date": str(event.start_date) if event and event.start_date else "",
            "end_date": str(event.end_date) if event and event.end_date else "",
            "theme": event.theme if event else "",
        },
        "attendance": {
            "registered_today": today_attendance is not None,
            "total_days": len(all_attendance),
            "records": [
                {"date": str(a.attendance_date.date() if a.attendance_date else a.created_at.date()), "id": a.id}
                for a in all_attendance
            ],
        },
    }


@router.post("/send_receipt/{registration_id}")
async def send_receipt(
    registration_id: int,
    body: SendReceiptSchema,
    background_tasks: BackgroundTasks,
    user: user_dependency,
    db: Session = Depends(get_db),
):
    """Generate a PDF receipt and email it to the participant."""
    registration = db.query(Registration).filter(Registration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    participant = registration.user
    if not participant:
        raise HTTPException(status_code=404, detail="Participant user not found")

    event = db.query(Event).filter(Event.id == registration.event_id).first()
    profile = participant.user_profile[0] if participant.user_profile else None

    # Amount paid — prefer Payment record, fall back to 0
    payment = registration.payment
    amount_paid = float(payment.payment_amount) if payment and payment.payment_amount else 0.0

    # Direct file URLs
    base_url = os.getenv("BASE_URL", "http://localhost:8001")
    payment_proof_url = None
    if registration.payment_proof:
        proof = registration.payment_proof
        # If it's already an http URL, keep as-is; otherwise build file URL
        if proof.startswith("http"):
            payment_proof_url = proof
        else:
            payment_proof_url = f"{base_url}/{proof}"

    photo_url = None
    photo_record = db.query(UserPhoto).filter(
        UserPhoto.user_id == participant.id,
        UserPhoto.deleted_at == None,
    ).order_by(UserPhoto.id.desc()).first()
    if photo_record and photo_record.path:
        if photo_record.path.startswith("http"):
            photo_url = photo_record.path
        else:
            photo_url = f"{base_url}/{photo_record.path}"

    pdf_bytes = generate_receipt_pdf(
        registration_id=registration.id,
        firstname=participant.firstname or "",
        lastname=participant.lastname or "",
        title_prefix=profile.title if profile else "",
        event_name=event.event if event else "Conference Registration",
        participation_role=registration.participation_role,
        amount_paid=amount_paid,
        date=datetime.utcnow(),
        membership_status=body.membership_status,
        membership_arrears=body.membership_arrears,
        payment_proof_url=payment_proof_url,
        photo_url=photo_url,
    )

    receipt_no = f"{registration.id}{datetime.utcnow().strftime('%m%d%Y')}"
    filename = f"Receipt_{receipt_no}.pdf"
    event_name_safe = (event.event or "Conference").replace(" ", "_")[:40]
    subject = f"Receipt – {event_name_safe}"

    full_name = " ".join(filter(None, [
        profile.title if profile else "",
        participant.firstname or "",
        participant.lastname or "",
    ]))
    html_body = f"""
    <div style="font-family:Arial,sans-serif;max-width:560px;margin:0 auto;">
      <div style="background:#fe5067;padding:16px 24px;">
        <h2 style="color:white;margin:0;font-size:18px;">ECSACONM – Payment Receipt</h2>
      </div>
      <div style="padding:24px;border:1px solid #eee;">
        <p>Dear <b>{full_name}</b>,</p>
        <p>Please find attached your official payment receipt for the <b>{event.event if event else 'conference'}</b>.</p>
        <p>Receipt No: <b style="color:#c0392b;">{receipt_no}</b></p>
        {"<p>Payment Proof: <a href='" + payment_proof_url + "'>View / Download</a></p>" if payment_proof_url else ""}
        {"<p>Badge Photo: <a href='" + photo_url + "'>View / Download</a></p>" if photo_url else ""}
        <hr style="border:none;border-top:1px solid #eee;margin:20px 0;">
        <p style="color:#888;font-size:12px;">© All Rights Reserved, ECSACONM</p>
      </div>
    </div>
    """

    background_tasks.add_task(
        send_email_with_attachment,
        participant.email,
        subject,
        html_body,
        pdf_bytes,
        filename,
    )

    return {"status": "sent", "receipt_no": receipt_no, "email": participant.email}


@router.get("/{event_id}")
async def get_event(
    request: Request,
    event_id: int,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    current_user: Optional[dict] = Depends(get_optional_current_user),
):
    client_ip = dependency.request_ip(request)

    dependency.log_activity(
        1,
        "VIEW_EVENT",
        "None",
        client_ip,
        f"View event id {event_id} and associated permissions",
    )

    if event := get_object(event_id, db, Event):
        registrations = event.registrations or []
        documents = event.documents or []
        links = event.links or []

        # ── Determine the current user's access level ─────────────────────────
        # "none"   → not logged in
        # "unpaid" → logged in but no paid registration for this event
        # "paid"   → has a paid registration OR has admin/secretariat permission
        user_access = "none"
        if current_user:
            uid = current_user["user_id"]
            # Check for admin permission (ADMIN_DASHBOARD) → always full access
            from models.models import UserRole, RolePermission, Permission as Perm
            is_admin = db.query(UserRole).join(
                RolePermission, UserRole.role_id == RolePermission.role_id
            ).join(
                Perm, RolePermission.permission_id == Perm.id
            ).filter(
                UserRole.user_id == uid,
                Perm.permission_code == "ADMIN_DASHBOARD",
            ).first()

            if is_admin:
                user_access = "paid"
            else:
                reg = db.query(Registration).filter(
                    Registration.user_id == uid,
                    Registration.event_id == event_id,
                    Registration.deleted_at == None,
                ).first()
                if reg and reg.paid:
                    user_access = "paid"
                else:
                    user_access = "unpaid"

        return {
            "event": {
                "id": event.id,
                "event": event.event,
                "country_id": event.country_id,
                "country": event.country.country if event.country else None,
                "org_unit_id": event.org_unit_id,
                "org_unit": event.org_unit.name if event.org_unit else None,
                "org_unit_primary_color": event.org_unit.primary_color if event.org_unit else "#0095B6",
                "org_unit_secondary_color": event.org_unit.secondary_color if event.org_unit else "#F7941D",
                "org_unit_logo": event.org_unit.logo if event.org_unit else None,
                "user_id": event.user_id,
                "firstname": event.user.firstname if event.user else None,
                "lastname": event.user.lastname if event.user else None,
                "location": event.location,
                "theme": event.theme,
                "description": event.description,
                "start_date": event.start_date,
                "end_date": event.end_date,
                "banner_image": event.banner_image,
                "organizers": event.organizers,
                "participation_info": event.participation_info,
                "logistics_info": event.logistics_info,
                "sponsors_info": event.sponsors_info,
                "participation_role": (
                    registrations[0].participation_role if registrations else None
                ),
                "user_access": user_access,
            },
            "participants": [
                {
                    "id": r.id,
                    "user_id": r.user_id,
                    "firstname": r.user.firstname if r.user else None,
                    "lastname": r.user.lastname if r.user else None,
                    "phone": r.user.phone if r.user else None,
                    "email": r.user.email if r.user else None,
                    "photo": (
                        r.user.user_photo[0].path
                        if r.user and r.user.user_photo and len(r.user.user_photo) > 0
                        else None
                    ),
                    "country_id": (
                        r.user.user_profile[0].country_id
                        if r.user and r.user.user_profile
                        else None
                    ),
                    "country": (
                        r.user.user_profile[0].country.country
                        if r.user and r.user.user_profile and r.user.user_profile[0].country
                        else None
                    ),
                    "title": (
                        r.user.user_profile[0].title
                        if r.user and r.user.user_profile
                        else None
                    ),
                    "designation": (
                        r.user.user_profile[0].designation
                        if r.user and r.user.user_profile
                        else None
                    ),
                    "participation_role": r.participation_role,
                    "organisation": (
                        r.user.user_profile[0].organisation
                        if r.user and r.user.user_profile
                        else None
                    ),
                    "paid": getattr(r, "paid", None),
                    "payment_proof": getattr(r, "payment_proof", None),
                    "payment_amount": (
                        float(r.payment.payment_amount)
                        if r.payment and r.payment.payment_amount else 0
                    ),
                    "registered_at": r.registered_at,
                }
                for r in registrations
            ],
            "documents": [
                {
                    "id": d.id,
                    "document_type": d.document_type,
                    "file_type": d.file_type,
                    "file_name": d.file_name,
                    "name": d.name,
                    "path": d.path,
                    "access_level": d.access_level,
                }
                for d in documents
            ],
            "links": [
                {
                    "id": l.id,
                    "name": l.name,
                    "link": l.link,
                    "access_level": l.access_level or "public",
                }
                for l in links
            ],
        }
    else:
        raise HTTPException(status_code=404, detail="event not found")


@router.put("/{event_id}")
async def update_event(
    request: Request,
    event_id: int,
    current_user: user_dependency,
    event_schema: EventUpdateSchema,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    auth_dependency.secure_access("UPDATE_EVENT", current_user["user_id"])

    client_ip = dependency.request_ip(request)

    dependency.log_activity(
        current_user["user_id"],
        "UPDATE_EVENT",
        current_user["username"],
        client_ip,
        f"Update event id {event_id}",
    )
    event_model = get_object(event_id, db, Event)

    event_model.org_unit_id = event_schema.org_unit_id
    event_model.country_id = event_schema.country_id
    event_model.user_id = current_user["user_id"]
    event_model.event = event_schema.event
    event_model.theme = event_schema.theme
    event_model.description = event_schema.description
    event_model.location = event_schema.location
    event_model.start_date = event_schema.start_date
    event_model.end_date = event_schema.end_date
    event_model.organizers = event_schema.organizers
    event_model.participation_info = event_schema.participation_info
    event_model.logistics_info = event_schema.logistics_info
    event_model.sponsors_info = event_schema.sponsors_info
    if event_schema.banner_image is not None:
        event_model.banner_image = event_schema.banner_image

    db.commit()
    db.refresh(event_model)
    return event_schema


@router.delete("/{event_id}")
async def delete_event(
    request: Request,
    event_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    auth_dependency.secure_access("DELETE_EVENT", current_user["user_id"])

    dependency.cascade_soft_delete_recursive(Event, event_id)

    client_ip = dependency.request_ip(request)
    event = get_object(event_id, db, Event)

    client_ip = dependency.request_ip(request)
    dependency.log_activity(
        current_user["user_id"],
        "DELETE_EVENT",
        current_user["username"],
        client_ip,
        f"Delete event id {event_id} event {event.event}",
    )
    return {"detail": "event Successfully deleted"}


@router.get("/registration/{registration_id}")
async def get_registration_details(
    request: Request,
    registration_id: int,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
):
    client_ip = dependency.request_ip(request)

    dependency.log_activity(
        1,
        "VIEW_REGISTRATION",
        "None",
        client_ip,
        f"View registration ID {registration_id}",
    )

    registration = get_object(registration_id, db, Registration)
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    user = db.query(User).filter(User.id == registration.user_id).first()
    event = db.query(Event).filter(Event.id == registration.event_id).first()

    return {
        "registration": {
            "id": registration.id,
            "user_id": registration.user_id,
            "event_id": registration.event_id,
            "participation_role": registration.participation_role.name,
            "paid": registration.paid,
            "created_at": registration.created_at,
            "updated_at": registration.updated_at,
        },
        "user": (
            {
                "id": user.id,
                "firstname": user.firstname,
                "lastname": user.lastname,
                "email": user.email,
                "phone": user.phone,
            }
            if user
            else None
        ),
        "event": (
            {
                "id": event.id,
                "event": event.event,
                "location": event.location,
                "start_date": event.start_date,
                "end_date": event.end_date,
            }
            if event
            else None
        ),
    }


@router.post("/registration/{user_id}")
async def event_registration(
    request: Request,
    user_id: int,
    registration_schema: RegistrationSchema,
    # current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    # auth_dependency: Auth = Depends(get_auth_dependency),
):
    # user = db.query(User).filter(User.id == user_id).first()
    # client_ip = dependency.request_ip(request)

    # dependency.log_activity(
    #     1,
    #     "EVENT_REGISTRATION",
    #     user.email,
    #     client_ip,
    #     f"Event ID: {registration_schema.event_id}",
    # )

    # Check for existing registration
    existing_registration = (
        db.query(Registration)
        .filter(
            Registration.user_id == user_id,
            Registration.event_id == registration_schema.event_id,
        )
        .first()
    )

    if existing_registration:
        # Update existing registration
        existing_registration.participation_role = (
            registration_schema.participation_role
        )
        db.commit()
        db.refresh(existing_registration)
        return {
            "message": "Registration updated successfully",
            "registration_id": existing_registration.id,
        }

    # Create new registration
    new_registration = Registration(
        user_id=user_id,
        event_id=registration_schema.event_id,
        participation_role=registration_schema.participation_role,
    )
    db.add(new_registration)
    db.commit()
    db.refresh(new_registration)

    return {
        "message": "Registration successful",
        "registration_id": new_registration.id,
    }


@router.delete("/registration/{event_id}")
async def event_deregistration(
    request: Request,
    event_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    client_ip = dependency.request_ip(request)

    dependency.log_activity(
        1,
        "EVENT_DEREGISTER",
        "None",
        client_ip,
        f"Deregister event id {event_id}",
    )

    existing_registration = (
        db.query(Registration)
        .filter(
            Registration.user_id == current_user["user_id"],
            Registration.event_id == event_id,
        )
        .first()
    )

    try:
        # First: delete any payment (if exists)
        existing_payment = (
            db.query(Payment)
            .filter(Payment.registration_id == existing_registration.id)
            .first()
        )
        if existing_payment:
            db.delete(existing_payment)
            db.flush()  # flush ensures DB sees the delete before the next delete

        # Now delete the registration
        db.delete(existing_registration)
        db.commit()

    except Exception as error:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete user event registration",
        ) from error


PAYMENT_RECEIPT_DIR = "uploads/payment_receipts"
if not os.path.exists(PAYMENT_RECEIPT_DIR):
    os.makedirs(PAYMENT_RECEIPT_DIR)


@router.delete("/deregister_participant/{registration_id}")
async def admin_deregister_participant(
    request: Request,
    registration_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Admin-only: deregister any participant by registration ID."""
    auth_dependency.secure_access("ADMIN_DASHBOARD", current_user["user_id"])

    registration = db.query(Registration).filter(Registration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    try:
        existing_payment = (
            db.query(Payment).filter(Payment.registration_id == registration.id).first()
        )
        if existing_payment:
            db.delete(existing_payment)
            db.flush()

        db.delete(registration)
        db.commit()
    except Exception as error:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to deregister participant",
        ) from error

    client_ip = dependency.request_ip(request)
    dependency.log_activity(
        current_user["user_id"],
        "ADMIN_DEREGISTER",
        current_user["username"],
        client_ip,
        f"Admin deregistered registration ID {registration_id}",
    )
    return {"detail": "Participant successfully deregistered"}


@router.post("/upload_payment_proof/{event_id}")
async def upload_payment_proof(
    event_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    file: UploadFile = File(None),
    payment_method: str = Form(default="bank_transfer"),
    payment_reference: str = Form(default=""),
    payment_amount: float = Form(default=0.0),
):
    """User uploads proof of payment for their registration."""
    from models.models import PaymentMethod as PM, PaymentStatus
    from datetime import datetime, timezone

    registration = db.query(Registration).filter(
        Registration.user_id == current_user["user_id"],
        Registration.event_id == event_id,
        Registration.deleted_at == None,
    ).first()

    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    try:
        file_path = None
        if file and file.filename:
            ext = os.path.splitext(file.filename)[1]
            unique_name = f"proof_{registration.id}_{uuid.uuid4().hex[:8]}{ext}"
            file_path = os.path.join(PAYMENT_RECEIPT_DIR, unique_name)
            with open(file_path, "wb+") as f:
                f.write(await file.read())
            registration.payment_proof = file_path

        # Map frontend payment_method string to enum
        method_map = {
            "bank_transfer": PM.BANK_TRANSFER,
            "mobile_money": PM.MPESA,
            "cash": PM.CASH,
            "card": PM.CARD,
        }
        method_enum = method_map.get(payment_method.lower(), PM.BANK_TRANSFER)

        # Create or update Payment record
        existing_payment = db.query(Payment).filter(
            Payment.registration_id == registration.id
        ).first()
        if existing_payment:
            existing_payment.payment_method = method_enum
            existing_payment.payment_reference = payment_reference or "N/A"
            existing_payment.payment_amount = payment_amount or 0
            existing_payment.payment_status = PaymentStatus.PENDING
            if file_path:
                existing_payment.payment_receipt = file_path
        else:
            new_payment = Payment(
                registration_id=registration.id,
                payment_date=datetime.now(timezone.utc),
                payment_method=method_enum,
                payment_reference=payment_reference or "N/A",
                payment_amount=payment_amount or 0,
                payment_status=PaymentStatus.PENDING,
                payment_receipt=file_path,
            )
            db.add(new_payment)

        db.commit()

        return JSONResponse(content={"status": "success", "payment_proof": file_path})
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/verify_payment/{registration_id}")
async def verify_payment(
    request: Request,
    registration_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Admin verifies a participant's payment, setting paid=True."""
    auth_dependency.secure_access("ADMIN_DASHBOARD", current_user["user_id"])

    registration = db.query(Registration).filter(Registration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    registration.paid = True
    db.commit()

    client_ip = dependency.request_ip(request)
    dependency.log_activity(
        current_user["user_id"],
        "VERIFY_PAYMENT",
        current_user["username"],
        client_ip,
        f"Verified payment for registration ID {registration_id}",
    )
    return {"detail": "Payment verified successfully"}


@router.put("/unverify_payment/{registration_id}")
async def unverify_payment(
    request: Request,
    registration_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Admin un-verifies a participant's payment, setting paid=False."""
    auth_dependency.secure_access("ADMIN_DASHBOARD", current_user["user_id"])

    registration = db.query(Registration).filter(Registration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    registration.paid = False
    db.commit()

    client_ip = dependency.request_ip(request)
    dependency.log_activity(
        current_user["user_id"],
        "UNVERIFY_PAYMENT",
        current_user["username"],
        client_ip,
        f"Un-verified payment for registration ID {registration_id}",
    )
    return {"detail": "Payment un-verified successfully"}


@router.post("/upload_document/")
async def upload_document(
    user: user_dependency,
    file: UploadFile = File(...),
    file_name: str = Form(...),
    doc_type: str = Form(...),
    access_level: str = Form(...),
    event_id: int = Form(...),
    db: Session = Depends(get_db),
):
    try:
        unique_dir = os.path.join(
            EVENT_DOCUMENT_DIR,
            f"{event_id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}",
        )
        os.makedirs(unique_dir, exist_ok=True)
        file_path = os.path.join(unique_dir, file.filename)
        with open(file_path, "wb+") as file_object:
            file_object.write(await file.read())

        event_document_model = Document(
            event_id=event_id,
            document_type=doc_type,
            file_type=file.content_type,
            file_name=file.filename,
            name=file_name,
            path=file_path,
            access_level=access_level,
        )
        db.add(event_document_model)
        db.commit()
        db.refresh(event_document_model)

        return JSONResponse(
            content={
                "status": "success",
                "message": f"File '{file.filename}' uploaded to '{unique_dir}'",
                "file_path": file_path,
            },
            status_code=200,
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")


@router.post("/upload_banner/{event_id}")
async def upload_banner(
    event_id: int,
    user: user_dependency,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    try:
        ext = os.path.splitext(file.filename)[1]
        unique_name = f"{event_id}_{uuid.uuid4().hex[:8]}{ext}"
        file_path = os.path.join(EVENT_BANNER_DIR, unique_name)
        with open(file_path, "wb+") as f:
            f.write(await file.read())

        event = db.query(Event).filter(Event.id == event_id).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        event.banner_image = file_path
        db.commit()

        return JSONResponse(content={"status": "success", "banner_image": file_path})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/delete_document/{document_id}")
async def delete_document(
    request: Request,
    document_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    auth_dependency.secure_access("DELETE_EVENT", current_user["user_id"])

    document = get_object(document_id, db, Document)

    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    document_folder = os.path.dirname(document.path)
    if os.path.exists(document_folder):
        try:
            shutil.rmtree(document_folder)
        except Exception as e:
            raise HTTPException(
                status_code=500, detail=f"Failed to delete folder: {str(e)}"
            )
    else:
        raise HTTPException(status_code=404, detail="Folder not found on disk")

    db.delete(document)
    db.commit()

    return {"status": "success", "message": f"Document and folder deleted successfully"}


@router.post("/add_link/")
async def add_link(
    request: Request,
    link_schema: LinkSchema,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    client_ip = dependency.request_ip(request)
    dependency.log_activity(
        current_user["user_id"],
        "ADD_EVENT",
        current_user["username"],
        client_ip,
        link_schema.event_id,
    )

    create_event_link_model = Link(
        event_id=link_schema.event_id,
        name=link_schema.name,
        link=str(link_schema.link),
        access_level=link_schema.access_level or "public",
    )

    db.add(create_event_link_model)
    db.commit()
    return link_schema


@router.put("/update_link/{link_id}")
async def update_link(
    request: Request,
    link_id: int,
    current_user: user_dependency,
    link_schema: LinkSchema,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    auth_dependency.secure_access("UPDATE_EVENT", current_user["user_id"])

    client_ip = dependency.request_ip(request)

    dependency.log_activity(
        current_user["user_id"],
        "UPDATE_EVENT",
        current_user["username"],
        client_ip,
        f"Update event id {link_id}",
    )
    link_model = get_object(link_id, db, Link)

    link_model.name = link_schema.name
    link_model.link = str(link_schema.link)

    db.commit()
    db.refresh(link_model)
    return link_schema


@router.delete("/delete_link/{link_id}")
async def delete_link(
    request: Request,
    link_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    auth_dependency.secure_access("DELETE_EVENT", current_user["user_id"])

    dependency.hard_delete(Link, link_id)

    client_ip = dependency.request_ip(request)
    link = get_object(link_id, db, Link)

    client_ip = dependency.request_ip(request)
    dependency.log_activity(
        current_user["user_id"],
        "DELETE_EVENT",
        current_user["username"],
        client_ip,
        f"Delete event id {link} event {link.name}",
    )
    return {"detail": "link Successfully deleted"}


# ── Bulk import participants from Excel ────────────────────────────────────────

# ── Google Drive download helper (used during bulk import) ─────────────────────
import re as _re, hashlib as _hashlib, mimetypes as _mimetypes, io as _io, time as _time

_DRIVE_ID_RE = _re.compile(r"[?&/]id=([a-zA-Z0-9_-]+)|/d/([a-zA-Z0-9_-]+)")

def _extract_drive_id(url: str):
    m = _DRIVE_ID_RE.search(url or "")
    if m:
        return m.group(1) or m.group(2)
    return None

def _drive_download(url: str, dest_dir: str, filename_stem: str):
    """Try to download a Google Drive file. Returns local relative path or None."""
    file_id = _extract_drive_id(url)
    if not file_id:
        return None
    token_path = os.path.join(os.path.dirname(__file__), "..", "scripts", "token.json")
    if not os.path.exists(token_path):
        return None   # no credentials cached — skip download, keep URL

    try:
        from google.oauth2.credentials import Credentials as GCreds
        from google.auth.transport.requests import Request as GRequest
        from googleapiclient.discovery import build as gbuild
        from googleapiclient.http import MediaIoBaseDownload

        creds = GCreds.from_authorized_user_file(token_path, ["https://www.googleapis.com/auth/drive.readonly"])
        if creds.expired and creds.refresh_token:
            creds.refresh(GRequest())

        service = gbuild("drive", "v3", credentials=creds)
        meta = service.files().get(fileId=file_id, fields="mimeType").execute()
        mime = meta.get("mimeType", "application/octet-stream")

        ext_map = {"image/jpeg": "jpg", "image/png": "png", "image/gif": "gif",
                   "image/webp": "webp", "application/pdf": "pdf"}
        ext = ext_map.get(mime) or (_mimetypes.guess_extension(mime) or ".bin").lstrip(".")

        request = service.files().get_media(fileId=file_id)
        buf = _io.BytesIO()
        dl = MediaIoBaseDownload(buf, request, chunksize=4 * 1024 * 1024)
        done = False
        while not done:
            _, done = dl.next_chunk()

        os.makedirs(dest_dir, exist_ok=True)
        uid = _hashlib.md5(file_id.encode()).hexdigest()[:8]
        fname = f"{filename_stem}_{uid}.{ext}"
        full_path = os.path.join(dest_dir, fname)
        with open(full_path, "wb") as fh:
            fh.write(buf.getvalue())

        # return a relative path (from api root)
        abs_api = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
        return os.path.relpath(full_path, abs_api)
    except Exception:
        return None   # fall back to storing the URL

# Maps Excel "Participation Category" text → ParticipationRole enum value
_ROLE_IMPORT_MAP = {
    "ecsacon members": "member_state",
    "ecsaconm members": "member_state",
    "member state": "member_state",
    "member_state": "member_state",
    "participant": "participant",
    "student": "student",
    "exhibitor": "exhibitor",
    "speaker": "speaker",
    "presenter": "presenter",
    "delegate": "delegate",
    "sponsor": "sponsor",
    "moderator": "moderator",
    "secretariat": "secretariat",
    "other africa": "other_africa",
    "international": "world",
    "moh": "moh",
    "ministry of health": "moh",
}


def _map_role(raw: str) -> ParticipationRole:
    key = (raw or "").strip().lower()
    value = _ROLE_IMPORT_MAP.get(key, "participant")
    return ParticipationRole(value)


def _normalize_header(h) -> str:
    return str(h or "").strip().lower()


def _detect_columns(headers: list) -> dict:
    """Return a dict mapping field name → column index (0-based)."""
    result = {}
    for i, h in enumerate(headers):
        n = _normalize_header(h)
        if "first" in n and "name" in n:
            result["firstname"] = i
        elif "last" in n and "name" in n:
            result["lastname"] = i
        elif n == "title":
            result["title"] = i
        elif "designation" in n:
            result["designation"] = i
        elif n == "country":
            result["country"] = i
        elif "email" in n:
            result.setdefault("email", i)           # first email column wins
        elif "telephone" in n or ("phone" in n and "code" not in n):
            result["phone"] = i
        elif "participation" in n or ("categor" in n and "membership" not in n):
            result["role"] = i
        elif "amount paid" in n:
            result["amount"] = i
        elif "full name" in n and ("certificate" in n or "appear" in n):
            result["certificate_name"] = i
        elif n == "address":
            result["address"] = i
        # Proof of payment — first occurrence wins
        elif "proof of payment" in n and "proof" not in result:
            result["proof"] = i
        # Badge / passport photo — first occurrence wins
        elif ("photo" in n or "badge" in n) and "photo" not in result:
            result["photo"] = i
        # Membership arrears status text (e.g. "Active member No Arrears")
        elif "membership arrears" in n and "arrears_status" not in result:
            result["arrears_status"] = i
        # Numeric arrears amount column (usually just labelled "arrears")
        elif n == "arrears" and "arrears_amount" not in result:
            result["arrears_amount"] = i
    return result


@router.post("/upload_participants/")
async def upload_participants(
    user: user_dependency,
    file: UploadFile = File(...),
    eventID: str = Form(...),
    db: Session = Depends(get_db),
):
    """Bulk-import participants from an Excel (.xlsx) or CSV file."""
    event_id = int(eventID)
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    contents = await file.read()

    # ── Parse the workbook ────────────────────────────────────────────────────
    try:
        wb = load_workbook(filename=BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    # Find header row — first row where we can detect at least email + firstname
    header_row_idx = 0
    col_map: dict = {}
    for idx, row in enumerate(rows):
        cm = _detect_columns(list(row))
        if "email" in cm and "firstname" in cm:
            header_row_idx = idx
            col_map = cm
            break

    if not col_map or "email" not in col_map:
        raise HTTPException(
            status_code=400,
            detail="Could not detect required columns. Expected: 'First Name', 'Last Name', 'Delegate Valid Email Address', etc.",
        )

    data_rows = rows[header_row_idx + 1:]

    created = updated = skipped = 0
    errors = []

    for row_num, row in enumerate(data_rows, start=header_row_idx + 2):
        def cell(field):
            idx = col_map.get(field)
            if idx is None:
                return None
            v = row[idx] if idx < len(row) else None
            return str(v).strip() if v is not None else None

        email = cell("email")
        if not email or "@" not in email:
            skipped += 1
            continue

        email = email.lower().strip()
        firstname = cell("firstname") or "Unknown"
        lastname = cell("lastname") or "Unknown"
        title_val = cell("title") or ""
        designation = cell("designation") or ""
        country_name = cell("country") or ""
        phone_raw = cell("phone") or ""
        role_raw = cell("role") or "participant"
        amount_raw = cell("amount")
        cert_name = cell("certificate_name") or ""

        # Sanitize phone: keep digits only, ensure non-empty
        phone_digits = "".join(c for c in phone_raw if c.isdigit())
        if not phone_digits:
            # Generate placeholder from email hash so it's unique
            import hashlib
            phone_digits = str(int(hashlib.md5(email.encode()).hexdigest()[:8], 16))[:12]

        # Resolve amount
        try:
            amount_paid = float(str(amount_raw).replace(",", "")) if amount_raw else 0.0
        except (ValueError, TypeError):
            amount_paid = 0.0

        # Resolve country
        country_obj = None
        if country_name:
            country_obj = (
                db.query(Country)
                .filter(Country.country.ilike(f"%{country_name}%"))
                .first()
            )

        # Resolve participation role
        try:
            role = _map_role(role_raw)
        except Exception:
            role = ParticipationRole.participant

        try:
            # ── Find or create User ───────────────────────────────────────────
            existing_user = db.query(User).filter(User.email == email).first()

            if existing_user:
                # Update name if blank
                if not existing_user.firstname or existing_user.firstname == "Unknown":
                    existing_user.firstname = firstname
                if not existing_user.lastname or existing_user.lastname == "Unknown":
                    existing_user.lastname = lastname
                user_obj = existing_user
                updated += 1
            else:
                # Check phone uniqueness — adjust if taken
                phone_candidate = phone_digits[:20]
                if db.query(User).filter(User.phone == phone_candidate).first():
                    import time
                    phone_candidate = phone_digits[:16] + str(int(time.time()))[-4:]

                user_obj = User(
                    firstname=firstname,
                    lastname=lastname,
                    email=email,
                    phone=phone_candidate,
                    hashed_password=hash_password("Ecsaconm@2025"),
                    verified=True,
                )
                db.add(user_obj)
                db.flush()  # get user_obj.id
                # Assign default "User" role (role_id=6)
                db.add(UserRole(user_id=user_obj.id, role_id=6))
                created += 1

            # ── Find or create UserProfile ────────────────────────────────────
            profile = db.query(UserProfile).filter(
                UserProfile.user_id == user_obj.id,
                UserProfile.deleted_at == None,
            ).first()

            if profile:
                if country_obj:
                    profile.country_id = country_obj.id
                if designation:
                    profile.designation = designation
                if title_val:
                    profile.title = title_val
                if cert_name:
                    profile.certificate_name = cert_name
            else:
                profile = UserProfile(
                    user_id=user_obj.id,
                    country_id=country_obj.id if country_obj else None,
                    title=title_val or '',
                    middle_name='',
                    gender='',
                    position='',
                    organisation='',
                    profession='',
                    designation=designation or '',
                    certificate_name=cert_name or '',
                    address='',
                )
                db.add(profile)

            # ── Find or create Registration ───────────────────────────────────
            reg = db.query(Registration).filter(
                Registration.user_id == user_obj.id,
                Registration.event_id == event_id,
            ).first()

            if reg:
                reg.participation_role = role
                if amount_paid > 0:
                    reg.paid = True
            else:
                reg = Registration(
                    user_id=user_obj.id,
                    event_id=event_id,
                    participation_role=role,
                    paid=amount_paid > 0,
                )
                db.add(reg)
                db.flush()

            # ── Create Payment record if paid ─────────────────────────────────
            if amount_paid > 0:
                reg.paid = True
                existing_payment = db.query(Payment).filter(
                    Payment.registration_id == reg.id
                ).first()
                if not existing_payment:
                    payment_rec = Payment(
                        registration_id=reg.id,
                        payment_date=datetime.utcnow(),
                        payment_method=PaymentMethod.BANK_TRANSFER,
                        payment_reference=f"IMPORT_{reg.id}",
                        payment_amount=amount_paid,
                        payment_status=PaymentStatus.COMPLETED,
                    )
                    db.add(payment_rec)
                else:
                    existing_payment.payment_amount = amount_paid
                    existing_payment.payment_status = PaymentStatus.COMPLETED

            # ── Save proof of payment (download Drive file if possible) ──────
            proof_url = cell("proof")
            if proof_url and proof_url.startswith("http"):
                local = _drive_download(
                    proof_url,
                    dest_dir="uploads/payment_receipts",
                    filename_stem=f"proof_{reg.id}",
                )
                reg.payment_proof = local if local else proof_url

            # ── Save address to profile ───────────────────────────────────────
            addr = cell("address")
            if addr and profile:
                profile.address = addr

            # ── Save badge photo (download Drive file if possible) ────────────
            photo_url = cell("photo")
            if photo_url and photo_url.startswith("http"):
                ts  = datetime.utcnow().strftime("%Y%m%d%H%M%S")
                uid = _hashlib.md5(f"{user_obj.id}{ts}".encode()).hexdigest()[:8]
                folder = f"uploads/picture/profile_picture/{user_obj.id}_{ts}_{uid}"
                local = _drive_download(
                    photo_url,
                    dest_dir=folder,
                    filename_stem="profile",
                )
                save_path = local if local else photo_url
                existing_photo = db.query(UserPhoto).filter(
                    UserPhoto.user_id == user_obj.id,
                    UserPhoto.deleted_at == None,
                ).first()
                if existing_photo:
                    existing_photo.path = save_path
                else:
                    db.add(UserPhoto(user_id=user_obj.id, path=save_path))

            db.commit()

        except Exception as e:
            db.rollback()
            errors.append(f"Row {row_num} ({email}): {str(e)}")

    return {
        "status": "complete",
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:10],  # return first 10 errors max
        "total_processed": created + updated + skipped,
    }


@router.get("/{event_id}/participants/download")
async def download_event_participants(
    request: Request,
    event_id: int,
    current_user: user_dependency,
    paid: Literal["all", "true", "false"] = Query("all"),
    db: Session = Depends(get_db),
    dependency=Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    # Permission check
    auth_dependency.secure_access("VIEW_EVENT", current_user["user_id"])
    client_ip = dependency.request_ip(request)

    dependency.log_activity(
        current_user["user_id"],
        "DOWNLOAD_PARTICIPANTS",
        current_user["username"],
        client_ip,
        f"Downloaded participants for event {event_id} with filter paid={paid}",
    )

    event = get_object(event_id, db, Event)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    # Collect participant details
    participants = []
    for reg in event.registrations:
        user = reg.user
        profile = user.user_profile[0] if user.user_profile else None
        country = profile.country.country if profile and profile.country else None
        organisation = profile.organisation if profile else None

        # Convert participation_role to string key (adjust this if ParticipationRole is Enum)
        role_key = (
            reg.participation_role.name
            if hasattr(reg.participation_role, "name")
            else str(reg.participation_role).lower()
        )

        participants.append(
            {
                "registration_id": reg.id,
                "firstname": user.firstname,
                "lastname": user.lastname,
                "email": user.email,
                "phone": user.phone,
                "title": profile.title if profile else "",
                "middle_name": profile.middle_name if profile else "",
                "gender": profile.gender if profile else "",
                "position": profile.position if profile else "",
                "organisation": organisation,
                "country": country,
                "participation_role": PARTICIPATION_ROLE_MAP.get(role_key, role_key),
                "paid": reg.paid,
                "registered_at": reg.registered_at,
            }
        )

    # Filter by paid status
    if paid != "all":
        is_paid = paid == "true"
        participants = [p for p in participants if p["paid"] == is_paid]

    if not participants:
        raise HTTPException(status_code=404, detail="No participants found")

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"

    headers = [
        "ID",
        "Title",
        "First Name",
        "Middle Name",
        "Last Name",
        "Gender",
        "Email",
        "Phone",
        "Organisation",
        "Position",
        "Country",
        "Participation Role",
        "Paid",
        "Registered At",
    ]
    ws.append(headers)

    for p in participants:
        ws.append(
            [
                p["registration_id"],
                p["title"],
                p["firstname"],
                p["middle_name"],
                p["lastname"],
                p["gender"],
                p["email"],
                p["phone"],
                p["organisation"] or "",
                p["position"],
                p["country"] or "",
                p["participation_role"],
                "Yes" if p["paid"] else "No",
                p["registered_at"].strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )

    # Stream response
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    safe_event_name = sanitize_filename(event.event)
    ascii_filename = f"{safe_event_name}_participants.xlsx"
    utf8_filename = urllib.parse.quote(ascii_filename)

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={ascii_filename}; filename*=UTF-8''{utf8_filename}"
        },
    )


@router.get("/with-registration/{user_id}")
@router.get("/with-registration/{user_id}/")
def list_events_with_user_registration(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),  # or your custom dependency
):
    # Get all events
    events = db.query(Event).all()

    # Get all registrations by the user
    user_regs = (
        db.query(Registration)
        .filter(Registration.user_id == user_id, Registration.deleted_at.is_(None))
        .all()
    )

    # Build a lookup dictionary for quick access: event_id -> registration
    reg_map = {reg.event_id: reg for reg in user_regs}

    # Construct result
    result = []
    for event in events:
        registration = reg_map.get(event.id)
        result.append(
            {
                "event": {
                    "id": event.id,
                    "title": event.event,
                    "theme": event.theme,
                    "description": event.description,
                    "start_date": event.start_date,
                    "end_date": event.end_date,
                    "location": event.location,
                    "country": event.country.country if event.country else None,
                    "org_unit": event.org_unit.name if event.org_unit else None,
                },
                "registered": bool(registration),
                "registration_details": (
                    {
                        "registration_id": registration.id,
                        "participation_role": registration.participation_role,
                        "paid": registration.paid,
                        "payment_status": "Paid" if registration.paid else "Not Paid",
                        "registered_at": registration.registered_at,
                    }
                    if registration
                    else None
                ),
            }
        )

    return result


def hex_to_rgb(hex_color: str):
    """Convert a hex color string like '#a02626' to a 0.0-1.0 RGB tuple."""
    hex_color = hex_color.lstrip("#")
    return tuple(int(hex_color[i: i + 2], 16) / 255.0 for i in (0, 2, 4))


def _render_badge_page(c, p, logo_left, logo_right, primary_rgb, secondary_rgb):
    """Draw a single badge page onto ReportLab canvas c."""
    width, height = (100 * mm, 140 * mm)

    # ── White background ─────────────────────────────────────────────────────
    c.setFillColorRGB(1, 1, 1)
    c.rect(0, 0, width, height, fill=True, stroke=False)

    # ── Crop marks ───────────────────────────────────────────────────────────
    mark_len = 5 * mm
    line_offset = 0.5 * mm
    c.setLineWidth(0.3)
    c.setStrokeColorRGB(0.5, 0.5, 0.5)
    c.line(0, height - line_offset, mark_len, height - line_offset)
    c.line(line_offset, height, line_offset, height - mark_len)
    c.line(width - mark_len, height - line_offset, width, height - line_offset)
    c.line(width - line_offset, height, width - line_offset, height - mark_len)
    c.line(0, line_offset, mark_len, line_offset)
    c.line(line_offset, 0, line_offset, mark_len)
    c.line(width - mark_len, line_offset, width, line_offset)
    c.line(width - line_offset, 0, width - line_offset, mark_len)

    # ── Colored top banner (primary_color) ───────────────────────────────────
    banner_h = 38 * mm
    c.setFillColorRGB(*primary_rgb)
    c.setStrokeColorRGB(*primary_rgb)
    c.rect(0, height - banner_h, width, banner_h, fill=True, stroke=False)

    # Logos inside the banner
    logo_size = 22 * mm
    c.drawImage(
        logo_left,
        4 * mm,
        height - logo_size - 4 * mm,
        logo_size,
        logo_size,
        preserveAspectRatio=True,
    )
    c.drawImage(
        logo_right,
        width - logo_size - 4 * mm,
        height - logo_size - 4 * mm,
        logo_size,
        logo_size,
        preserveAspectRatio=True,
    )

    # Event name in white on the banner
    event_name_y = height - banner_h + 5 * mm
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(width / 2, event_name_y, normalize_event_name(p["event_name"]))

    # ── Participant details ───────────────────────────────────────────────────
    c.setFillColorRGB(0, 0, 0)
    c.setStrokeColorRGB(0, 0, 0)

    y = height - banner_h - 10 * mm
    full_name = (
        f"{p['title']} {p['firstname']} {p['middle_name']} {p['lastname']}".strip()
    )
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(width / 2, y, full_name)

    if p["position"]:
        y -= 8 * mm
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(width / 2, y, p["position"])

    if p["organisation"]:
        y -= 7 * mm
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(width / 2, y, p["organisation"])

    if p["country"]:
        y -= 7 * mm
        c.setFont("Helvetica", 10)
        c.drawCentredString(width / 2, y, p["country"])

    # Participant ID
    y -= 7 * mm
    c.setFont("Helvetica", 9)
    c.drawCentredString(width / 2, y, f"Participant ID: BPF{p['registration_id']}")

    if p["location"]:
        y -= 7 * mm
        c.setFont("Helvetica-Oblique", 9)
        c.drawCentredString(width / 2, y, f"📍 {p['location']}")

    # ── Colored role strip (secondary_color) ─────────────────────────────────
    role_strip_h = 9 * mm
    role_strip_y = 33 * mm
    c.setFillColorRGB(*secondary_rgb)
    c.setStrokeColorRGB(*secondary_rgb)
    c.rect(0, role_strip_y, width, role_strip_h, fill=True, stroke=False)
    if p["participation_role"]:
        c.setFillColorRGB(1, 1, 1)
        c.setFont("Helvetica-Bold", 9)
        c.drawCentredString(
            width / 2, role_strip_y + 2.5 * mm, p["participation_role"].upper()
        )

    # ── QR code (links to attendance check-in) ───────────────────────────────
    qr_size = 28 * mm
    qr_y = 3.5 * mm
    qr_data = (
        f"{CLIENT_ORIGIN}/event-attendance/{p['event_id']}?reg={p['registration_id']}"
    )
    qr = qrcode.make(qr_data)
    qr_buf = BytesIO()
    qr.save(qr_buf, format="PNG")
    qr_buf.seek(0)
    c.drawImage(ImageReader(qr_buf), (width - qr_size) / 2, qr_y, qr_size, qr_size)

    c.showPage()


@router.get("/{event_id}/participants/badges")
async def download_participant_badges_pdf(
    request: Request,
    event_id: int,
    current_user: user_dependency,
    paid: Literal["all", "true", "false"] = Query("all"),
    db: Session = Depends(get_db),
    dependency=Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    auth_dependency.secure_access("VIEW_EVENT", current_user["user_id"])
    client_ip = dependency.request_ip(request)
    dependency.log_activity(
        current_user["user_id"],
        "DOWNLOAD_BADGES",
        current_user["username"],
        client_ip,
        f"Downloaded participant badges for event {event_id} with filter paid={paid}",
    )

    event = get_object(event_id, db, Event)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    primary_color = (event.org_unit.primary_color or "#0095B6") if event.org_unit else "#0095B6"
    secondary_color = (event.org_unit.secondary_color or "#F7941D") if event.org_unit else "#F7941D"
    primary_rgb = hex_to_rgb(primary_color)
    secondary_rgb = hex_to_rgb(secondary_color)

    participants = []
    for reg in event.registrations:
        user = reg.user
        profile = user.user_profile[0] if user.user_profile else None
        country = profile.country.country if profile and profile.country else None
        organisation = profile.organisation if profile else None
        role_key = (
            reg.participation_role.name
            if hasattr(reg.participation_role, "name")
            else str(reg.participation_role).lower()
        )
        participants.append(
            {
                "registration_id": reg.id,
                "event_id": event_id,
                "title": profile.title if profile else "",
                "firstname": user.firstname,
                "middle_name": profile.middle_name if profile else "",
                "lastname": user.lastname,
                "position": profile.position if profile else "",
                "organisation": organisation,
                "country": country,
                "participation_role": PARTICIPATION_ROLE_MAP.get(role_key, role_key),
                "event_name": event.event,
                "location": event.location or "",
                "paid": reg.paid,
            }
        )

    if paid != "all":
        is_paid = paid == "true"
        participants = [p for p in participants if p["paid"] == is_paid]

    if not participants:
        raise HTTPException(status_code=404, detail="No participants found")

    buffer = BytesIO()
    width, height = (100 * mm, 140 * mm)
    c = canvas.Canvas(buffer, pagesize=(width, height))

    logo_left = convert_png_to_rgb("assets/logo_left.png")
    logo_right = convert_png_to_rgb("assets/logo_right.png")

    for p in participants:
        _render_badge_page(c, p, logo_left, logo_right, primary_rgb, secondary_rgb)

    c.save()
    buffer.seek(0)

    safe_event_name = sanitize_filename(event.event)
    ascii_filename = f"{safe_event_name}_participant_badges.pdf"
    utf8_filename = urllib.parse.quote(ascii_filename)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={ascii_filename}; filename*=UTF-8''{utf8_filename}"
        },
    )


@router.get("/{event_id}/my-badge")
async def download_my_badge(
    event_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    """Generate a badge PDF for the currently authenticated paid user."""
    event = get_object(event_id, db, Event)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    reg = db.query(Registration).filter(
        Registration.user_id == current_user["user_id"],
        Registration.event_id == event_id,
        Registration.deleted_at == None,
    ).first()

    if not reg:
        raise HTTPException(status_code=404, detail="You are not registered for this event")
    if not reg.paid:
        raise HTTPException(status_code=403, detail="Badge is only available after payment is confirmed")

    user = reg.user
    profile = user.user_profile[0] if user.user_profile else None
    country = profile.country.country if profile and profile.country else None
    organisation = profile.organisation if profile else None
    role_key = (
        reg.participation_role.name
        if hasattr(reg.participation_role, "name")
        else str(reg.participation_role).lower()
    )

    primary_color = (event.org_unit.primary_color or "#0095B6") if event.org_unit else "#0095B6"
    secondary_color = (event.org_unit.secondary_color or "#F7941D") if event.org_unit else "#F7941D"
    primary_rgb = hex_to_rgb(primary_color)
    secondary_rgb = hex_to_rgb(secondary_color)

    p = {
        "registration_id": reg.id,
        "event_id": event_id,
        "title": profile.title if profile else "",
        "firstname": user.firstname,
        "middle_name": profile.middle_name if profile else "",
        "lastname": user.lastname,
        "position": profile.position if profile else "",
        "organisation": organisation,
        "country": country,
        "participation_role": PARTICIPATION_ROLE_MAP.get(role_key, role_key),
        "event_name": event.event,
        "location": event.location or "",
        "paid": reg.paid,
    }

    buffer = BytesIO()
    width, height = (100 * mm, 140 * mm)
    c = canvas.Canvas(buffer, pagesize=(width, height))

    logo_left = convert_png_to_rgb("assets/logo_left.png")
    logo_right = convert_png_to_rgb("assets/logo_right.png")

    _render_badge_page(c, p, logo_left, logo_right, primary_rgb, secondary_rgb)

    c.save()
    buffer.seek(0)

    safe_name = sanitize_filename(f"{user.firstname}_{user.lastname}")
    ascii_filename = f"badge_{safe_name}.pdf"
    utf8_filename = urllib.parse.quote(ascii_filename)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={ascii_filename}; filename*=UTF-8''{utf8_filename}"
        },
    )


@router.get("/{event_id}/attendance")
async def get_event_attendance(
    event_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Admin: get all attendance records for an event."""
    auth_dependency.secure_access("VIEW_EVENT", current_user["user_id"])

    from models.models import EventAttendance
    records = (
        db.query(EventAttendance)
        .join(Registration, EventAttendance.registration_id == Registration.id)
        .filter(Registration.event_id == event_id)
        .order_by(EventAttendance.attendance_date.desc())
        .all()
    )

    result = []
    for a in records:
        reg = a.registration
        user = reg.user if reg else None
        profile = user.user_profile[0] if user and user.user_profile else None
        result.append({
            "id": a.id,
            "registration_id": a.registration_id,
            "attendance_date": a.attendance_date,
            "firstname": user.firstname if user else "",
            "lastname": user.lastname if user else "",
            "email": user.email if user else "",
            "organisation": profile.organisation if profile else "",
            "country": profile.country.country if profile and profile.country else "",
            "participation_role": reg.participation_role.name if reg else "",
            "paid": reg.paid if reg else False,
        })

    return {"total": len(result), "data": result}
